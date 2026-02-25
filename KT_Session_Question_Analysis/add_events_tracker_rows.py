#!/usr/bin/env python3
"""Add Events session (2026-02-11) Q&A rows to SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx."""
from pathlib import Path
import openpyxl

DIR = Path(__file__).resolve().parent
WB_PATH = DIR / "SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx"
SESSION_DATE = "2026-02-11"

ROWS = [
    ("View as published shows changes fast but publisher takes too long; new job (CF) not showing on publisher. Can we expedite?", SESSION_DATE, "Gonzalo Calasich (SHRSS)", "Need more info (cache, etc.). Daniela will monitor during call. Andy investigating failed publish logs; will raise support ticket.", SESSION_DATE, "Daniela Tea / Andy Lambert", "Pending", "Gonzalo: not showing at all on publisher; likely cache."),
    ("Override for date field (e.g. 'every Wednesday through February') instead of actual date/time?", SESSION_DATE, "Lisa Cardia", "Not in event content fragment. Promotions has date override; can confirm if same behavior wanted for events—gap.", SESSION_DATE, "Daniela Tea", "Deferred", "Same question for location override."),
    ("What happens if we add an additional image in the event image field (multi)?", SESSION_DATE, "Lisa Cardia", "Believes it defaults to first one; can test. Migration used only one image.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Event status: logic so presale dates show presale disclaimer then auto-revert to 'on sale'?", SESSION_DATE, "Lisa Cardia", "Not currently. Gap: associate status with date/time so it changes without author republish. Presale can be outside business hours.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Hide CTA (more info) when there's no landing page / no extra info for user?", SESSION_DATE, "Lisa Cardia", "Default links to event detail page. Toggle to hide CTA entirely would be gap.", SESSION_DATE, "Daniela Tea", "Deferred", "Use case: weekly free entertainment; no detail page."),
    ("Location reference: limit by user so cafe only adds their location? Avoid choosing wrong location (e.g. Amsterdam for Mexico event).", SESSION_DATE, "Lisa Cardia", "Location reference drives filter roll-up; root path limits which events show. Wrong location in CF still adds to this site's filter. Restrict by user = gap.", SESSION_DATE, "Daniela Tea", "Deferred", "Folder path = organization; location reference = required for display/filter."),
    ("Default image for event card so author doesn't have to select every time (generic default)?", SESSION_DATE, "Lisa Cardia", "CF cannot be saved without image. Component-level default (like jobs listing) could be gap. Author can use default from DAM in meantime.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Event detail page—why does it look like a card? Are we on careers? Hard Rock Live—can we use for careers?", SESSION_DATE, "Mayte Eme", "This is the event detail page (one template). Same for all sites. Daniela created new page to show from scratch.", SESSION_DATE, "Daniela Tea", "Answered", "Mayte: digital page not working for careers events."),
    ("Clean/friendly URL for event detail (property.com/artist-name not numbers)?", SESSION_DATE, "Lisa Cardia / Mayte Eme", "Event ID is read-only to prevent duplicates. Enhancement = make editable for friendly URL. Mayte: need real names; like WordPress/Sitecore—pick up page name. Gap.", SESSION_DATE, "Daniela Tea", "Deferred", "Labor intensive; add to gap list."),
    ("Event detail page hero—standard across all or customizable per page?", SESSION_DATE, "Edwin Aquino", "Default hero on event detail page; banner image + alt in CF overrides.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Event statuses: all manual or can we schedule (presale, announcement, rescheduled)? Do they drive/hide content?", SESSION_DATE, "Mayte Eme", "Currently manual. Gap to have date range per status. Presale/sold out etc. change label/flag; event status message is separate. Doesn't hide CTA or content.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Do we have to publish to see changes? How to preview before publish for approval?", SESSION_DATE, "Lisa Cardia / Mayte Eme", "View as published doesn't show CF changes until CF is published. Andy: preview tier exists, not configured; will check export. Mayte: need browser preview to screenshot for approval.", SESSION_DATE, "Daniela Tea / Andy Lambert", "Pending", "Preview server / preview links = gap analysis."),
    ("When adding new event: publish fragment, detail page, and calendar every time?", SESSION_DATE, "Lisa Cardia", "One event detail page per calendar—publish once. New event: publish CF; calendar page must be published for event to show on publisher (can take a few minutes).", SESSION_DATE, "Daniela Tea", "Answered", "Gonzalo's delay = calendar page publish."),
    ("Unpublish event (take down without delete) to repurpose later?", SESSION_DATE, "Lisa Cardia / Mayte Eme", "Unpublish CF removes from calendar. Daniela to check if detail URL still accessible. Later: unpublished = detail URL shows configured 'no events' message; content not visible.", SESSION_DATE, "Daniela Tea", "Answered", "Mayte: add as gap—not expected behavior but at least user doesn't see it."),
    ("Recap: unpublish works; schedule not local to property; end date auto drop-off; all statuses manual; no date/location override; three times instead of one (vs Sitecore).", SESSION_DATE, "Mayte Eme", "Daniela acknowledged. Hiring events use promotions model (has overrides); events model does not.", SESSION_DATE, "Mayte Eme / Daniela Tea", "Answered", ""),
    ("Hiring events—using promotions not events? What can events CF do vs promotions?", SESSION_DATE, "Mayte Eme / Gonzalo", "Hiring events section uses promotions content fragment (date/location override). Events CF + calendar/detail = what we covered. Tomorrow: hiring events component.", SESSION_DATE, "Daniela Tea", "Answered", "Careers not using events CF for that section."),
    ("Can we choose mix-and-match locations for one calendar (e.g. six properties, or hotel+casinos, or by venue)?", SESSION_DATE, "Mayte Eme", "No. One content fragment folder path only; cannot select multiple folders. Big gap for multi-venue/entertainment sites.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Pull events into other pages (venue page, homepage by category, e.g. dinner and a show)?", SESSION_DATE, "Mayte Eme", "Content Fragment card list component can show events (or news); by root path or by tag (e.g. dinner and a show). Not locked to calendar component.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Why did we lose the filter for events (careers hiring events landing page)?", SESSION_DATE, "Lisa Cardia", "Careers uses content fragment card (no filter on purpose for homepage). Promotions component has filter; can swap on landing page if desired.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Can we pull jobs (hot jobs/hiring events) to property page (e.g. Reverb) by tag or location?", SESSION_DATE, "Mayte Eme", "Yes—select root path (e.g. Tampa folder); all jobs under that path. Cannot select multiple locations in one component. Exclude category under Tampa? Not in component.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Display tag on event cards (21+, free event, phone-free)?", SESSION_DATE, "Edwin Aquino", "Daniela asked for example; will discuss with technical team.", SESSION_DATE, "Daniela Tea", "Pending", "Edwin to send example."),
    ("Preview link for unpublished page to share for approval before republish?", SESSION_DATE, "Rick Lyon (Director of Digital Experience)", "Unpublished CF not on calendar or detail. If publish CF, visible in author; page must be published for end user. Daniela to report back on detail-page behavior.", SESSION_DATE, "Daniela Tea", "Pending", "Preview before publish = gap."),
    ("Publish delays on stage—investigating?", SESSION_DATE, "Lucas Nelson", "Andy: seen failed publish entries in cloud services log; raising support ticket; will keep team posted.", SESSION_DATE, "Andy Lambert", "Pending", ""),
    ("Governance: Who owns Event Categories tags and DPLT location list? Process when property/venue added or renamed?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Authoring: One-page checklist for event authoring (what to create, publish, order) and when to use events CF vs promotions CF?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Tagging: Event categories vs tags for filtering vs 'dinner and a show' query—who governs?", SESSION_DATE, "Product Director", "Not answered in session. Categories = filter; tags = internal; tag for query possible.", "", "", "Pending", "Product Director persona."),
    ("Gap analysis: Decision criteria for events CF vs promotions CF; which gaps apply to which model?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Permissions: Who can create/edit Event Categories tags and DPLT location reference options?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb["Events"]
    next_row = ws.max_row + 1
    for row in ROWS:
        for col, val in enumerate(row, 1):
            ws.cell(row=next_row, column=col, value=val)
        next_row += 1
    wb.save(WB_PATH)
    print(f"Added {len(ROWS)} rows to Events sheet.")

if __name__ == "__main__":
    main()
