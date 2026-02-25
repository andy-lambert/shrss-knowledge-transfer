#!/usr/bin/env python3
"""Add Careers session (2026-02-12) Q&A rows to SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx."""
from pathlib import Path
import openpyxl

DIR = Path(__file__).resolve().parent
WB_PATH = DIR / "SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx"
SESSION_DATE = "2026-02-12"

# Tuple order: Question/Comment, Date Asked, Asked By, Answer, Answered On, Answered By, Status, Notes
# Workbook columns: Status, Date Asked, Answered On, Answered By, Asked By, Question/Comment, Answer, Notes
NEW_COL_ORDER = (7, 2, 5, 6, 3, 1, 4, 8)
ROWS = [
    ("Manual job content fragment: create, publish, not showing on jobs page. Other CFs show. Workday sync deletes manual jobs (is not API data) in ~30 min.", SESSION_DATE, "Gonzalo Calasich (SHRSS)", "Andy: investigating; support ticket. Get all job IDs not excluding manual; email with recommendation coming; tech sync to drill in.", SESSION_DATE, "Andy Lambert / Daniela Tea", "Pending", "Gonzalo to send example URLs for logs."),
    ("Promotion ID—where from? What do we put? Is it made up? Required for targeting/analytics we had for casino, not unique ID for display.", SESSION_DATE, "Lisa Cardia / Mayte Eme", "Author-entered unique identifier to display one promotion (e.g. content fragment card). No format requirement. Uniqueness validated on save. Same value in another CF blocks save.", SESSION_DATE, "Daniela Tea", "Answered", "Mayte: gap—we don't want this; need to remember/copy; not smart (dropdown of existing IDs)."),
    ("When pasting ID elsewhere (e.g. component), is there a dropdown to select from existing or must we remember/paste?", SESSION_DATE, "Mayte Eme", "Component field is plain text; no dropdown. Must copy from CF. Gonzalo: gap = source field to navigate DAM and pick CF.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Without promotions filter (careers homepage): steps to get card on page = know ID, copy from fragment, paste in component. No other way?", SESSION_DATE, "Lisa Cardia", "Yes. Content fragment card expects promotion ID for promotions/events. No folder path picker. Gap: enhance card list to include promotions or add path/source picker.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Promotion ID for casino or careers? We're not using for casinos; we need to fix for casinos (many users, can't remember IDs).", SESSION_DATE, "Mayte Eme", "Initially for casinos. Careers using it for hiring events. Lucas: capture what it doesn't have; disposition in platform expansion; modify/change is part of it.", SESSION_DATE, "Daniela Tea / Lucas Nelson", "Deferred", ""),
    ("Promotion search: can we remove the search field and keep only filters?", SESSION_DATE, "Mayte Eme", "Search cannot be removed currently. With filters, search bar always shows. To display as list without filter/search = gap.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("If we use promotions for hiring events we can't have promotion cards in grid/carousel elsewhere without filter+search—only one-by-one by ID.", SESSION_DATE, "Mayte Eme", "Correct. Content fragment card list currently events/news only; gap = add promotions so list/carousel without filter+search.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Who manages filter options (category, venue, date range)? Author or dev? Can we add different options?", SESSION_DATE, "Lisa Cardia", "In component code; hardcoded from promotions requirements. New filter = dev. Potential to move to generic list (author-managed) but not available today. Mayte: we manage every filter—gap.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Promotion search: same as events—can't select two category folders (e.g. casino + hotel promotions)?", SESSION_DATE, "Lisa Cardia", "Single path only. Can choose promotions root or e.g. casinos; cannot pick two of four folders. Mayte: another gap.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Promotions: are filter options (e.g. select category) translatable / overridable for language?", SESSION_DATE, "Lisa Cardia", "Daniela to check with tech team. Components have language strings; need to confirm what's translatable. At minimum override all areas. Gonzalo: include Transperfect + developer labels.", SESSION_DATE, "Daniela Tea", "Pending", "Technical topic; Scott/technical enablement."),
    ("Transperfect: manual translations for some languages; not always Transperfect. OK?", SESSION_DATE, "Mayte Eme", "Noted. Daniela will address in Transperfect KT.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Promotion start/end date and time—is that for scheduling or display only?", SESSION_DATE, "Mayte Eme", "Display only. Schedule = Manage Publication > Now or Later (activation date). Not for 'display this event on Saturday but publish tomorrow.'", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("If we schedule for later and then edit the fragment (e.g. change venue), does it maintain schedule or push live?", SESSION_DATE, "Lisa Cardia", "Daniela to test (save, manage publication, save); will confirm.", SESSION_DATE, "Daniela Tea", "Pending", ""),
    ("Promo types (loyalty, slot)—hardcoded? How do we add extra promo types?", SESSION_DATE, "Lisa Cardia", "In content fragment model (not component). Model editor = more flexibility; permissions apply. Adding dropdown to component = dev. Model edit vs component = different.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Card/promotion image: 16:9 asset—how to control focal point (e.g. center guitar)? Image position?", SESSION_DATE, "Don Middlebrook", "Image position tab on image components (desktop/tablet/mobile). List (promotion search) has no per-row position; crop outside or use rendition—DAM/gap.", SESSION_DATE, "Daniela Tea", "Answered", "Don: one source image across placements; figure out with DAM/renditions."),
    ("Job listings on another site (e.g. hotel)—same component, theme changes styling? Can we choose multiple locations?", SESSION_DATE, "Mayte Eme", "Yes—component is global; theme drives look. Still one root path only (gap). Use case: Tampa/Hollywood dedicated job pages.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("How do we see the promotion card's detail page? Is there a promotion detail template?", SESSION_DATE, "Lisa Cardia", "No promotion detail template. Create open page, add content, link from CF. Daniela to show casino example from integration env.", SESSION_DATE, "Daniela Tea", "Pending", "Banner image in CF = for when promotion has its own page."),
    ("Events template allows link to another page (event detail). Promotions can't do that—we have hot jobs/dealer university with detail pages.", SESSION_DATE, "Mayte Eme", "Promotions: can link to any page from CF. Events: one detail page + ID. Override in events = gap; then could use events for hiring events. Workaround: link to existing page from promotion CF.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Video card: external URL—recommended width? Pixel required? What does width affect?", SESSION_DATE, "Lisa Cardia", "Width affects modal size. Number only, no 'px.' 846 was to match live site. Can change.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Video card: third party (YouTube) gives fixed/responsive, aspect ratio; external URL doesn't. Same config for external?", SESSION_DATE, "Lisa Cardia / Mayte Eme", "External URL = no those options. Third party = full config. Gap—external should have same controls; else only use YouTube/Vimeo. Mayte: don't use external until fixed.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Video card styling (none/black/white): is 'none' transparent? White description not showing.", SESSION_DATE, "Lisa Cardia", "Daniela to check white; none = default. Black works. Screenshot of three when fixed.", SESSION_DATE, "Daniela Tea", "Pending", ""),
    ("Video card: if we add a fourth, does it become carousel automatically (like Sitecore)?", SESSION_DATE, "Mayte Eme", "No. Video card not in allowed components for card carousel or hero carousel. Gap: allow video cards in carousel.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Card (icon variation): with 6 or 9 cards, does it go 3 to 1 columns automatically or manual per breakpoint?", SESSION_DATE, "Mayte Eme", "Manual per viewport for standalone card. Some components (e.g. carousel) have breakpoint config. Mayte: add to gap list.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Executives/diversity page (careers) same as hardrock.com—shared so one update?", SESSION_DATE, "Mayte Eme", "Currently separate cards. Can move to experience fragment: update once, same styling both sites. Content fragment = same content, different styling per site.", SESSION_DATE, "Daniela Tea", "Answered", "Rick: corporate page hardrock.com."),
    ("Experience fragment: same content, different styling per site possible?", SESSION_DATE, "Mayte Eme", "Experience fragment = same content + same look. For same content, different look = content fragment.", SESSION_DATE, "Daniela Tea", "Answered", ""),
    ("Asset specs for video card thumbnail, promotion card/banner—recommended dimensions? Focal point for video thumbnail?", SESSION_DATE, "Lisa Cardia", "Document on component basis; video thumbnail aspect/specs; no image tab on video card. Gap.", SESSION_DATE, "Daniela Tea", "Deferred", ""),
    ("Store component on page but not displayed (like Sitecore disable)—reuse later without recreating? Hide = still in DOM?", SESSION_DATE, "Lisa Cardia / Mayte Eme", "Hide component (layout > eye) = still in tree, still in DOM/crawled. Not like Sitecore where disabled = not in HTML. Gap: store without rendering for evergreen swap.", SESSION_DATE, "Daniela Tea", "Deferred", "Content refresh initiative; rotate temp vs evergreen."),
    ("Careers locations page same as brand—pulling from DPLT? Can we copy accordion/component so one source?", SESSION_DATE, "Mayte Eme", "Careers = accordion/text (migrated); brand = map component + DPLT. Can copy map component to careers. Intent: all location listings from DPLT, one maintenance.", SESSION_DATE, "Daniela Tea", "Answered", "Daniela to post test page with copied component."),
    ("Governance: Who owns promotion ID list and filter options? Process to add/change?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Authoring: Single checklist for hiring events (CF → card vs search → detail page → publish order) and when events vs promotions?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Tagging: Do promotions use tags for filtering? If card list adds promotions, tag-based selection?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Gap priorities: Which items (ID vs path, filter removal, card list promotions, video carousel, thumbnail specs, hide vs disable) for careers vs platform expansion?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Permissions: Who can edit promotions model (promo types) and promotion search config (filter labels)?", SESSION_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb["Careers"]
    next_row = ws.max_row + 1
    for row in ROWS:
        ordered = tuple(row[i - 1] for i in NEW_COL_ORDER)
        for col, val in enumerate(ordered, 1):
            ws.cell(row=next_row, column=col, value=val)
        next_row += 1
    wb.save(WB_PATH)
    print(f"Added {len(ROWS)} rows to Careers sheet.")

if __name__ == "__main__":
    main()
