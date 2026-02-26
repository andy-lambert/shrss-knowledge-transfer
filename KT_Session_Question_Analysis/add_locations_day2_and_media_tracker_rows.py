#!/usr/bin/env python3
"""Add Locations Day 2 (2026-02-24) and Media (2026-02-25) sheets and tracker rows."""
from pathlib import Path
import openpyxl

DIR = Path(__file__).resolve().parent
WB_PATH = DIR / "SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx"
NEW_HEADERS = ("Status", "Date Asked", "Answered On", "Answered By", "Asked By", "Question/Comment", "Answer", "Notes")
# Tuple order in data: Question/Comment, Date Asked, Asked By, Answer, Answered On, Answered By, Status, Notes
NEW_COL_ORDER = (7, 2, 5, 6, 3, 1, 4, 8)

LOC_DAY2_DATE = "2026-02-24"
MEDIA_DATE = "2026-02-25"

LOCATIONS_DAY2_ROWS = [
    ("What path do we choose for CF list path (category listing)? Not intuitive; can we update info tools?", LOC_DAY2_DATE, "Edwin Aquino", "Path = parent; component shows children of that parent. E.g. path to EN under news → year folders. Daniela demonstrated; Edwin confirmed.", LOC_DAY2_DATE, "Daniela Tea", "Answered", "News session follow-up."),
    ("Delivery widget: Hard Rock Casino Rockford showing—casino classified as cafe. Who do we contact to fix?", LOC_DAY2_DATE, "Lisa Cardia", "DPLT team (Vipul Patel). Classifications come from DPLT; if misclassified, take to Vipul's team. Lisa to reach out.", LOC_DAY2_DATE, "Kerry Holyoak (SHRSS) / Daniela Tea", "Answered", "Line of business = Cafe in DPLT drives delivery list."),
    ("Booking modal: black border; can we make booking widget same size/background as modal?", LOC_DAY2_DATE, "Rick Lyon", "Note as gap—configurable color/size for modal.", LOC_DAY2_DATE, "Daniela Tea", "Deferred", ""),
    ("Where do we get destination value (e.g. 59391) for booking widget? Do authors need to know it?", LOC_DAY2_DATE, "Lisa Cardia", "Rick: Senex/Vizergy; in booking links. Migrated; values don't change. Authors may need for restore if someone changes. Daniela: casino booking not finalized (pre-pause).", LOC_DAY2_DATE, "Rick Lyon / Daniela Tea", "Answered", ""),
    ("Booking widget: adding a different field (e.g. balcony) beyond rooms/adults/children—development?", LOC_DAY2_DATE, "Lisa Cardia", "Yes—new field type, UI, parameter passing, booking engine side. Development for sure.", LOC_DAY2_DATE, "Daniela Tea", "Answered", ""),
    ("Booking offers: can we hide the booking window section completely?", LOC_DAY2_DATE, "Rick Lyon", "Tried; window still present, fields blank. Hide label too = potential gap. No option to hide completely today.", LOC_DAY2_DATE, "Daniela Tea", "Deferred", ""),
    ("Find a location: can we change the background color?", LOC_DAY2_DATE, "Lisa Cardia", "Currently fixed in code. Change would require code update to allow author-configurable background.", LOC_DAY2_DATE, "Daniela Tea", "Deferred", ""),
    ("Google Map: regions dropdown blank in prod—why? North America first in config but showing last.", LOC_DAY2_DATE, "Gonzalo Calasich (SHRSS)", "Permissions: groups didn't have read on generic list 'regions.' Template Authors given read. Order: ensure regions authored in dialog (first = top); republish page.", LOC_DAY2_DATE, "Daniela Tea / Lucas Nelson", "Answered", "ACS Commons on prod. Media session confirmed fix."),
    ("Google Map: filter map not in config but filters showing—why? No results vs no locations?", LOC_DAY2_DATE, "Lisa Cardia", "Filter map adds filters on map. Different messages for no results vs no locations. Daniela made copy to demo filter map.", LOC_DAY2_DATE, "Daniela Tea", "Answered", ""),
    ("Google Map: carousel dots not centered under images.", LOC_DAY2_DATE, "Lisa Cardia / Rick Lyon", "Daniela noted; would need to be updated to center align carousel dots.", LOC_DAY2_DATE, "Daniela Tea", "Pending", ""),
    ("Who owns DPLT classification (line of business) and SHRSS contact for misclassification?", "2026-02-24", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Which groups need read/write on generic list 'regions' for Google Map?", "2026-02-24", "Product Director", "Not answered in session. Andy covering permissions in KT next week.", "", "", "Pending", "Product Director persona."),
]

MEDIA_ROWS = [
    ("Google Map regions dropdown not visible yesterday—resolution?", MEDIA_DATE, "Gonzalo Calasich (SHRSS)", "Generic list 'regions' populates dropdown. User groups lacked read access. Template Authors given read; Gonzalo and Lisa confirmed can see. Permissions topic with Andy next week.", MEDIA_DATE, "Daniela Tea", "Answered", "Locations Day 2 follow-up."),
    ("North America showing last on prod—publish page again?", MEDIA_DATE, "Gonzalo Calasich (SHRSS)", "View as published to confirm order; then publish. Andy to take a look at prod page.", MEDIA_DATE, "Daniela Tea", "Answered", ""),
    ("Media gallery: carousel dots not centered (left of center).", MEDIA_DATE, "Rick Lyon", "Daniela noted; capture for update to center align carousel dots.", MEDIA_DATE, "Daniela Tea", "Pending", ""),
    ("Image gallery grid: option to show/hide title or inherit from asset like alt text?", MEDIA_DATE, "Edwin Aquino", "Not for this component. Title pulled from DAM metadata. If no title in metadata, 'no title' shows. Enhancement to capture if team wants option to not display title.", MEDIA_DATE, "Daniela Tea", "Deferred", ""),
    ("Image gallery grid: different titles for same image per property/site?", MEDIA_DATE, "Edwin Aquino", "Not currently. Control via metadata per asset; no variation/title override in component.", MEDIA_DATE, "Daniela Tea", "Deferred", ""),
    ("Image gallery grid: are we adding titles/descriptions to assets?", MEDIA_DATE, "Daniela Tea (to Don)", "Don: planning to add as much metadata as possible—titles, description. Daniela: if titles not always desired, capture enhancement to not display or use description.", MEDIA_DATE, "Don Middlebrook / Daniela Tea", "Answered", ""),
    ("Who can add/edit generic list values (e.g. regions)? Which groups get read vs write?", MEDIA_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Image gallery grid: hide title or show description—enhancement?", MEDIA_DATE, "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

def add_sheet_with_rows(wb, sheet_name, rows):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start = ws.max_row + 1
    else:
        ws = wb.create_sheet(sheet_name)
        for c, h in enumerate(NEW_HEADERS, 1):
            ws.cell(row=1, column=c, value=h)
        start = 2
    for row in rows:
        ordered = tuple(row[i - 1] for i in NEW_COL_ORDER)
        for c, val in enumerate(ordered, 1):
            ws.cell(row=start, column=c, value=val)
        start += 1

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    add_sheet_with_rows(wb, "Locations_Day_2", LOCATIONS_DAY2_ROWS)
    add_sheet_with_rows(wb, "Media", MEDIA_ROWS)
    wb.save(WB_PATH)
    print(f"Locations_Day_2: {len(LOCATIONS_DAY2_ROWS)} rows; Media: {len(MEDIA_ROWS)} rows.")

if __name__ == "__main__":
    main()
