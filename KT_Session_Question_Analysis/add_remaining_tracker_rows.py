#!/usr/bin/env python3
"""Add Tagging, DAM, Shared Data, News, Locations session rows to SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx."""
from pathlib import Path
import openpyxl

DIR = Path(__file__).resolve().parent
WB_PATH = DIR / "SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx"

def add_rows(ws, rows):
    next_row = ws.max_row + 1
    for row in rows:
        for col, val in enumerate(row, 1):
            ws.cell(row=next_row, column=col, value=val)
        next_row += 1

# Tagging_Taxonomy_Metadata_Gov — 2026-02-17
TAGGING = [
    ("Why multiple category directories (categories, category, news categories, events categories)? Consolidate under one?", "2026-02-17", "Don Middlebrook", "Different purposes; naming could be clearer. Andy to assess functional implications of consolidation, then green-light. Add as to-do for cleanup.", "2026-02-17", "Andy Lambert", "Pending", "Don: same for property/hotel/casino/cafe under locations."),
    ("If we restructure tags, does that break anything currently?", "2026-02-17", "Lisa Cardia", "Assess first; goal is no hard-coded paths, config where possible. Then SHRSS can make changes. Move tool updates reference paths.", "2026-02-17", "Andy Lambert", "Pending", ""),
    ("How do we use generic lists (ACS Commons path-to-tag mapping)? What are they used for?", "2026-02-17", "Don Middlebrook", "Governs tags for a section. Andy to add to Confluence Q&A; Vinay for deeper answer. Not used for corporate/careers today.", "2026-02-17", "Andy Lambert", "Pending", ""),
    ("Content fragment model category field—tag-driven or defined in model? Where do these lists come from?", "2026-02-17", "Don Middlebrook", "Can be tag-driven or enumeration. Locations example = not all from tags. Depends on use case; gap to align. Check model.", "2026-02-17", "Andy Lambert", "Deferred", ""),
    ("Asset categories vs content fragment categories—different? Event categories on asset vs events.", "2026-02-17", "Don Middlebrook", "Yes—asset metadata vs CF categories; confusing. Asset event categories ≠ CF event categories. Action: cleanup, rename for clarity.", "2026-02-17", "Andy Lambert", "Pending", "Lisa: location reference ID same issue (events session)."),
    ("Pre-populated metadata set for export (so we don't add 15 properties every time)?", "2026-02-17", "Don Middlebrook", "Not OOTB. Andy: pain; keep text file with property names, copy/paste. Chris may say same tomorrow.", "2026-02-17", "Andy Lambert", "Deferred", ""),
    ("Governance: Who owns SHRSS tag vocabulary and consolidation decision? Process to add namespace?", "2026-02-17", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Authoring: Where do authors see tag-driven vs fixed list? One-pager by model/asset type?", "2026-02-17", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Permissions: Which groups have create/update/delete on SHRSS tags?", "2026-02-17", "Product Director", "Not answered in session. Andy: use groups; lock down.", "", "", "Pending", "Product Director persona."),
]

# DAM_Training_Usage_Admin — 2026-02-18
DAM = [
    ("SHRSS primary: consolidate into SHRSS, deprecate primary? Path forward.", "2026-02-18", "Don Middlebrook", "Yes. Don to move primary content into SHRSS; deprecate primary. Content fragment behavior may need back-end look.", "2026-02-18", "Andy Lambert", "Answered", "Don: 13 months ago didn't know move wouldn't break refs."),
    ("Legacy/migrated assets folder—option to move old content there, new structure for new?", "2026-02-18", "Andy Lambert", "Yes—e.g. legacy-assets; new content in new structure; cleanup/scripting later. Chris agreed.", "2026-02-18", "Andy Lambert / Chris Lewis", "Answered", ""),
    ("Language folders for assets—required for translation automation?", "2026-02-18", "Andy Lambert", "Chris: language at top level for AM Guide etc. If no language-swap use case, metadata for language is fine. Don: single asset, metadata.", "2026-02-18", "Chris Lewis", "Answered", ""),
    ("Renditions: Don tried to create, they're not applying.", "2026-02-18", "Don Middlebrook", "Andy: renditions in agenda; will cover. Hot topic.", "2026-02-18", "Andy Lambert", "Pending", ""),
    ("Careers asset names odd (C2D5 etc.)—not from Sitecore. Bulk rename?", "2026-02-18", "Lisa Cardia / Don Middlebrook", "Chris: migration tool internal ref. Bulk rename via export metadata to Excel, edit names, re-upload.", "2026-02-18", "Chris Lewis", "Answered", ""),
    ("Governance: Who approves DAM folder structure and bulk renames?", "2026-02-18", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
    ("Rendition profiles not applying—root cause and fix in scope?", "2026-02-18", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

# Shared_Data — 2026-02-19
SHARED = [
    ("Redirect vanity URL checkbox—what does it do? When to use?", "2026-02-19", "Edwin Aquino", "Daniela to get back. In this case not needed.", "2026-02-19", "Daniela Tea", "Pending", ""),
    ("Multiple vanity URLs per page—all link to same page?", "2026-02-19", "Rick Lyon", "Yes. Useful for print (short URL) vs long.", "2026-02-19", "Daniela Tea", "Answered", ""),
    ("Card carousel tablet field removed—why? Can we have tablet-specific again?", "2026-02-19", "Daniela Tea (update)", "Requirement: desktop/tablet same, mobile separate. Tablet always 2 for non–full width; field had no effect. Removed. Gap if tablet config needed.", "2026-02-19", "Daniela Tea", "Deferred", ""),
    ("Content fragment model: default image for new CF? Override by author?", "2026-02-19", "Edwin Aquino", "Yes—default value on model; new CFs get it; author can replace. Custom component (e.g. card) would show it.", "2026-02-19", "Daniela Tea", "Answered", ""),
    ("Experience fragment variation: break connection to reorder or use as regular component?", "2026-02-19", "Edwin Aquino", "Daniela to confirm.", "2026-02-19", "Daniela Tea", "Pending", ""),
    ("Governance: Who creates CF models for shared data? Who creates EF variations and rolls out?", "2026-02-19", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

# News — 2026-02-20
NEWS = [
    ("List component: exclude certain tags (e.g. do not use) from list?", "2026-02-20", "Edwin Aquino", "Not OOTB; matching only. Daniela to take use case back (e.g. promotion page not on landing).", "2026-02-20", "Daniela Tea", "Pending", ""),
    ("List: what dictates which items appear in which column? Order?", "2026-02-20", "Edwin Aquino", "Order by title or last modified. Column count in config; fills in order.", "2026-02-20", "Daniela Tea", "Answered", ""),
    ("Logo styling (image icon in footer)—what is it? Use case?", "2026-02-20", "Edwin Aquino", "Daniela to check JIRA; not footer-specific; image component in general (e.g. height/width for multi-logo pages).", "2026-02-20", "Daniela Tea", "Pending", ""),
    ("Redirect vanity URL checkbox: 302 redirect; when to use?", "2026-02-20", "Daniela Tea (update)", "Documentation: 302 when checkbox. Daniela to share link; technical enablement for redirect management.", "2026-02-20", "Daniela Tea", "Answered", ""),
    ("News: default image by category/tag (e.g. casino vs hotel)?", "2026-02-20", "Edwin Aquino", "Not currently. Gap; need requirements (multiple tags = which default?).", "2026-02-20", "Daniela Tea", "Deferred", ""),
    ("News CF: default image in model? Carlos: all articles have image.", "2026-02-20", "Carlos Aldana", "Default image at component level (news search results) for items without image. Model default can be added.", "2026-02-20", "Daniela Tea", "Answered", ""),
    ("Governance: Who can add news categories (tags) and change news CF model?", "2026-02-20", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

# Locations — 2026-02-23
LOCATIONS = [
    ("Location CF title = location ID; how to bulk update to property legal name?", "2026-02-23", "Don Middlebrook", "Bulk metadata update: take value from property legal name (or chosen field), replace title. Daniela: tech KT if help needed.", "2026-02-23", "Daniela Tea", "Answered", "Read-only fields from DPLT; title editable."),
    ("New location from DPLT: is delivery only—author manually checks?", "2026-02-23", "Lisa Cardia", "Yes.", "2026-02-23", "Daniela Tea", "Answered", ""),
    ("Image in location CF—where does it populate? Hotels?", "2026-02-23", "Lisa Cardia", "Destinations and venues component (e.g. Hard Rock Hotel Cancun). Image specs to Confluence.", "2026-02-23", "Daniela Tea", "Answered", ""),
    ("Delivery/venue links: open in new window?", "2026-02-23", "Rick Lyon", "Yes; baked in (external links).", "2026-02-23", "Daniela Tea", "Answered", ""),
    ("Add new sort-by or filter field (e.g. meeting room view)? Process?", "2026-02-23", "Lisa Cardia", "Update venue CF model, add to component dialog, front-end component update. Mayte: gap—overcomplicated; need optimized way. Lucas: mark as gap.", "2026-02-23", "Daniela Tea", "Deferred", ""),
    ("Filters (e.g. type of destination) also hard-coded? Can we add (e.g. Unity participating hotel)?", "2026-02-23", "Lisa Cardia", "Same concept—model + component.", "2026-02-23", "Daniela Tea", "Deferred", ""),
    ("Regions filter on hardrock.com prod not working—Google Map component?", "2026-02-23", "Gonzalo Calasich (SHRSS)", "Different component (Google Map). Destination search and filter = this one. Cover Google Map when we get to it.", "2026-02-23", "Daniela Tea", "Pending", ""),
    ("Gap list: Adobe building it or SHRSS? Confluence questions = gap list?", "2026-02-23", "Lisa Cardia / Mayte Eme", "Lucas: SHRSS documents gaps; Confluence questions feed funnel. Adobe shows framework; SHRSS comes to platform expansion with use cases and gaps.", "2026-02-23", "Lucas Nelson", "Answered", "Scott: questions/needs on Confluence."),
    ("Governance: Who maintains location/venue CF when DPLT updates? Who can add new sort/filter fields?", "2026-02-23", "Product Director", "Not answered in session.", "", "", "Pending", "Product Director persona."),
]

def main():
    wb = openpyxl.load_workbook(WB_PATH)
    add_rows(wb["Tagging_Taxonomy_Metadata_Gov"], TAGGING)
    add_rows(wb["DAM_Training_Usage_Admin"], DAM)
    add_rows(wb["Shared_Data"], SHARED)
    add_rows(wb["News"], NEWS)
    add_rows(wb["Locations"], LOCATIONS)
    wb.save(WB_PATH)
    print(f"Tagging: {len(TAGGING)}; DAM: {len(DAM)}; Shared_Data: {len(SHARED)}; News: {len(NEWS)}; Locations: {len(LOCATIONS)}")

if __name__ == "__main__":
    main()
