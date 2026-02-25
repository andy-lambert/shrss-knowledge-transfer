#!/usr/bin/env python3
"""Convert markdown tables in all .md files in this directory to one Excel workbook."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

DIR = Path(__file__).resolve().parent

def parse_md_table(path):
    """Parse first markdown table in file. Returns list of rows (each row is list of cells)."""
    text = path.read_text(encoding="utf-8")
    rows = []
    in_table = False
    for line in text.splitlines():
        line = line.rstrip()
        if not line.startswith("|"):
            if in_table:
                break
            continue
        in_table = True
        # Split by |, strip each cell, drop first/last empty from leading/trailing |
        cells = [c.strip() for c in line.split("|")]
        if not cells:
            continue
        # Remove first and last if empty (from leading/trailing pipe)
        if cells and cells[0] == "":
            cells = cells[1:]
        if cells and cells[-1] == "":
            cells = cells[:-1]
        # Skip separator row (--- or :--- etc.)
        if cells and all(
            set(c.replace(":", "").strip()) <= set("-") for c in cells
        ):
            continue
        rows.append(cells)
    return rows


def main():
    md_files = sorted(DIR.glob("*.md"))
    if not md_files:
        print("No .md files found")
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for md_path in md_files:
        rows = parse_md_table(md_path)
        if not rows:
            continue
        # Sheet name: from filename, max 31 chars, no invalid chars
        base = md_path.stem.replace("SHRSS_KT_Session_Questions_", "").replace("_", " ")[:31]
        for c in ["*", "?", ":", "\\", "/", "[", "]"]:
            base = base.replace(c, "")
        if not base:
            base = md_path.stem[:31]
        ws = wb.create_sheet(title=base)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(bold=True)
    out = DIR / "KT_Session_Question_Analysis.xlsx"
    wb.save(out)
    print(f"Wrote {out} with {len(wb.sheetnames)} sheets")


if __name__ == "__main__":
    main()
