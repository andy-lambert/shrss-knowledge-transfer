#!/usr/bin/env python3
"""
Analyze KT questions for likely AI-generation.
Populates AI-Generated (TRUE/FALSE), Confidence (1-100), Reasoning.
Cross-references with consolidated transcript; uses task criteria (formal phrasing,
technical buzzwords, reworded snippet, duplicate feel, irrelevant to topic).
Heuristic-based; thresholds and patterns can be tuned. Re-run to refresh all rows.
"""
import re
from pathlib import Path

import openpyxl

DIR = Path(__file__).resolve().parent
TRANSCRIPT_PATH = DIR.parent / "KT_Session_Transcripts" / "SHRSS_Adobe_KT_All_Session_Transcripts_Consolidated.md"
WB_PATH = DIR / "SHRSS_KT_Session_Questions.xlsx"
SHEET = "All_SHRSS_KT_Questions"

# Workbook Session value -> transcript section heading prefix (first line of section)
SESSION_TO_HEADING = {
    "Jobs": "Session: Jobs —",
    "Events": "Session: Events —",
    "Careers": "Session: Careers —",
    "Tagging_Taxonomy_Metadata_Gov": "Session: Tagging & Taxonomy —",
    "DAM_Training_Usage_Admin": "Session: DAM —",
    "Shared_Data": "Session: Shared Data —",
    "News": "Session: News —",
    "Locations": "Session: Locations",  # matches both Locations and Locations, Day 2
}

# Strong signals for likely AI-generated (technical / formal)
TECHNICAL_TERMS = re.compile(
    r"\b(OSGi|Sling\s*model|query\s*builder|search\s*index|GraphQL|JCR|"
    r"CF-driven|XF-driven|indexation|indexed\s*differently|LD\s*JSON|JSON-LD|"
    r"versioning\s*strategy|taxonomy\s*architecture|namespace\s*\(|Sling\s*Model|"
    r"content\s*fragment\s*model|experience\s*fragment|persistent\s*query|"
    r"generic\s*list|path-to-tag|metadata\s*schema|rendition|processing\s*profile)\b",
    re.I
)
FORMAL_PHRASES = re.compile(
    r"\b(To what extent|What is the current intended|What is the versioning strategy|"
    r"How do we dynamically|Is there any scenario where|Could you clarify|Please advise|"
    r"It would be helpful to understand|I would like to know|With regard to|With respect to|"
    r"the extent to which|the degree to which|differentiator|methodology|best practice|"
    r"desired outcome|intended behavior|expected behavior|source of truth|single source of truth|"
    r"taxonomy\s*architecture|intended\s*taxonomy|namespace\s*under|"
    r"query\s*and\s*display|indexed\s*differently|dynamically\s*query|outside of the\s*component|"
    r"publish\s*cycles|multiple\s*sites|configuration\s*options|role-based|"
    r"production-ready|interim\s*solutions|restricted\s*to|component\s*only)\b",
    re.I
)
# Conversational / human-like (lowers AI score)
CONVERSATIONAL = re.compile(
    r"\b(can we|do we|is there|are we|will we|can I|do I|how do I|where do I|"
    r"what happens if|why does|when do|who can|who is|quick question|"
    r"just wondering|sorry to ask|follow-up|follow up)\b",
    re.I
)

def load_transcript_sections(path: Path) -> dict:
    """Return dict: section_key -> full section text (no heading)."""
    text = path.read_text(encoding="utf-8", errors="replace")
    sections = {}
    current_key = None
    current_lines = []
    for line in text.splitlines():
        if line.startswith("## Session:"):
            if current_key is not None:
                sections[current_key] = "\n".join(current_lines)
            current_lines = []
            # Derive key from heading
            if "Jobs —" in line:
                current_key = "Jobs"
            elif "Events —" in line:
                current_key = "Events"
            elif "Careers —" in line:
                current_key = "Careers"
            elif "Tagging & Taxonomy —" in line:
                current_key = "Tagging_Taxonomy_Metadata_Gov"
            elif "DAM —" in line:
                current_key = "DAM_Training_Usage_Admin"
            elif "Shared Data —" in line:
                current_key = "Shared_Data"
            elif "News —" in line:
                current_key = "News"
            elif "Locations, Day 2 —" in line:
                current_key = "Locations_Day2"
            elif "Locations —" in line:
                current_key = "Locations"
            elif "Media —" in line:
                current_key = "Media"
            else:
                current_key = "Other"
        else:
            current_lines.append(line)
    if current_key is not None:
        sections[current_key] = "\n".join(current_lines)
    # Merge Locations and Locations_Day2 for "Locations" session
    if "Locations" in sections and "Locations_Day2" in sections:
        sections["Locations"] = (sections.get("Locations", "") + "\n" + sections["Locations_Day2"])
    return sections


def score_question(question: str, transcript: str) -> tuple[bool, int, str]:
    """
    Returns (is_likely_ai: bool, confidence: 1-100, reasoning: str).
    """
    if not question or not isinstance(question, str):
        return False, 10, "Empty or invalid question; default FALSE."
    q = question.strip()
    q_lower = q.lower()
    word_count = len(q.split())
    reasons = []
    ai_score = 0  # 0–100 scale; higher = more likely AI

    # Length: very long often AI
    if word_count > 80:
        ai_score += 28
        reasons.append("very long, formal question")
    elif word_count > 50:
        ai_score += 15
        reasons.append("long question")
    elif word_count > 30:
        ai_score += 5
        reasons.append("moderate length")

    # Technical terms
    tech_matches = TECHNICAL_TERMS.findall(q)
    if tech_matches:
        ai_score += min(30, 10 + 5 * len(tech_matches))
        reasons.append(f"technical/AEM terms: {', '.join(list(set(tech_matches))[:5])}")

    # Formal/AI phrasing
    formal_matches = FORMAL_PHRASES.findall(q)
    if formal_matches:
        ai_score += min(35, 15 + 5 * len(formal_matches))
        reasons.append(f"formal/AI phrasing: {', '.join(list(set(formal_matches))[:3])}")

    # Conversational: lowers score (strong human signal)
    if CONVERSATIONAL.search(q):
        ai_score -= 18
        reasons.append("conversational phrasing (human-like)")
    # Baseline: questions with no strong human signal and 20+ words often reworded/formal
    elif word_count >= 20 and not CONVERSATIONAL.search(q):
        ai_score += 8
        reasons.append("moderate length without conversational cues")

    # Check if question looks like a reworded transcript snippet (key words close in transcript)
    if transcript:
        # Extract significant words (alpha, len > 2, not common)
        stop = {"the", "and", "for", "are", "but", "not", "you", "all", "can", "had", "her", "his", "was", "one", "our", "out", "has", "how", "its", "may", "now", "any", "did", "get", "got", "let", "new", "see", "way", "who", "with", "this", "that", "from", "what", "when", "where", "which", "will", "your", "about", "into", "than", "them", "then", "they", "would", "could", "should", "there", "their", "have", "been", "being", "does", "other", "only", "some", "such", "than", "these", "were", "what", "while", "would"}
        words = set(re.findall(r"[a-z]{3,}", q_lower)) - stop
        if len(words) >= 3:
            # See if several of these appear close together in transcript (within 400 chars)
            transcript_lower = transcript.lower()
            for i in range(0, len(transcript_lower) - 500, 250):
                chunk = transcript_lower[i : i + 500]
                count = sum(1 for w in words if w in chunk)
                if count >= min(3, len(words)):
                    ai_score += 22
                    reasons.append("question appears to rephrase a transcript snippet")
                    break

    # Clamp and decide (threshold: task states majority of 543 are noise / likely AI-generated)
    ai_score = max(0, min(100, ai_score))
    is_ai = ai_score >= 22
    if is_ai:
        confidence = max(50, min(95, ai_score + 5))
        reasoning = "Likely AI-generated: " + "; ".join(reasons) if reasons else "Formal/technical phrasing and structure."
    else:
        confidence = max(5, 100 - ai_score - 10)
        reasoning = "Likely human: " + ("; ".join(reasons) if reasons else "conversational or short, no strong AI markers.")

    return is_ai, confidence, reasoning


def main():
    sections = load_transcript_sections(TRANSCRIPT_PATH)
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb[SHEET]
    updated = 0
    for row in range(2, ws.max_row + 1):
        session = ws.cell(row=row, column=2).value
        question = ws.cell(row=row, column=3).value
        transcript = ""
        if session:
            if session == "Locations":
                transcript = sections.get("Locations", "") or ""
            else:
                transcript = sections.get(session, "") or ""
        is_ai, confidence, reasoning = score_question(question, transcript)
        ws.cell(row=row, column=4, value="TRUE" if is_ai else "FALSE")
        ws.cell(row=row, column=5, value=f"{confidence}%")
        ws.cell(row=row, column=6, value=reasoning[:500] if reasoning else "")  # cap reasoning length
        updated += 1
    wb.save(WB_PATH)
    print(f"Updated {updated} rows with AI-Generated, Confidence, Reasoning.")


if __name__ == "__main__":
    main()
