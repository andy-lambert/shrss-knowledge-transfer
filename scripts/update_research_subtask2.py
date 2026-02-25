#!/usr/bin/env python3
"""
Subtask 2: Populate Answer and Answered By (Adobe) for questions that can be
confidently answered from KT transcripts and implementation docs.
Skips: declarative sentences (no ?), questions starting with "Why".
"""
import re
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "KT_Session_Follow_Up_Questions_RESEARCH_WORKING.xlsx"

# Answers derived from KT transcript (Jobs session) and implementation docs.
# Key = substring of question (lowercase) to match; value = (answer_text, optional_row_hint).
# First matching key wins; keep keys specific.
ANSWER_MAP = [
    # Jobs - from transcript
    (
        "based on demo, seems we can use job components on other sites",
        "Yes. Job components (e.g. Job Search, Job Listing) are reusable and can be added to any page that uses a template that allows them. The demo showed adding the Job Search component to a new page.",
    ),
    (
        "are filters driven dynamically from workday data or statically configured",
        "Filters are driven dynamically from the job content fragment data, which is populated by the Workday sync. Filter options (job category, job type, properties, locations) are derived from the content fragments currently in AEM; they are not hard-coded. Sort options (e.g. most recent, job title, location) are configured on the Job Search component.",
    ),
    (
        "if \"is api data\" is unchecked",
        "When \"Is API data\" is unchecked, the content fragment is treated as author-created (not from Workday). Workday-synced fields will not overwrite that fragment on the next sync. Author-only fields (image, is hot job, LD JSON) are never overwritten by sync regardless.",
    ),
    (
        "the address from api, where is that used",
        "The address (country, state, city) from the API is used in the job search results display and to build the filter \"Select locations\" (combined field). It also dictates the content fragment folder structure in AEM (e.g. country / state / city / job postings).",
    ),
    (
        "what happens if a job detail page url exists but job no longer exists in workday",
        "When a job is removed from Workday, it is no longer part of the sync payload. On the next sync, that job’s content fragment (and thus the job detail experience) is removed from AEM. The folder structure may remain, but the posting content is gone.",
    ),
    (
        "what is the finalized dam folder structure for job-related images",
        "Job content fragments are organized under a folder structure by country, state, and city (e.g. US > FL > Tampa). Job-related images can be set via the image configuration page (by location) or overridden per fragment in the CF image field; the exact DAM path for that configuration is implementation-specific.",
    ),
    (
        "what happens if a category/location/property/type has no jobs in filters",
        "Filter options in the Job Search component are built from the jobs currently in AEM (e.g. \"Select properties\" shows only properties that have existing results). If a category/location/property/type has no jobs, it typically would not appear as a filter option or would show with zero results if the UI allows selecting it.",
    ),
    (
        "does the override replace the image we see on the card and the listing or just the card",
        "If you override the image on the content fragment, it updates the image used in both the job listings (cards) and the job search component (longer horizontal result cards).",
    ),
    (
        "fields that get overwritten by sync",
        "Only fields that come from Workday get overwritten on sync. The image field, Is hot job toggle, and LD JSON field do not come from Workday; authors can set these and they will not be overwritten by the sync.",
    ),
    (
        "can we use job components on other sites",
        "Yes. Job components can be used on any page that supports them (e.g. Careers and other sites); they query job content fragments in AEM.",
    ),
    (
        "how is sequence determined",
        "Sequence/sort is configured on the Job Search component (e.g. most recent, job title, location). Default sort options are provided in the component config; hot jobs stay at the top, and the rest sort according to the selected option.",
    ),
    (
        "do hot jobs automatically drop off",
        "Hot jobs are controlled by the \"Is hot job\" toggle on the content fragment, which is author-maintained and not overwritten by Workday. There is no automatic drop-off by end date mentioned in the KT; scheduling of content fragments (activation date) is available for author-created content.",
    ),
    (
        "when job is removed from workday",
        "When a job is removed or no longer active in Workday, the next sync run will remove that job’s content fragment from AEM. The folder structure (e.g. country/state/city) remains; only the job postings come and go.",
    ),
    (
        "can an author create a job",
        "Yes. Create > Content Fragment > select the Jobs model. This will be an author-created job (Is API data = off). Category and other Workday-sourced fields are not dropdowns; you must enter values that match what you want (e.g. exact category text).",
    ),
    (
        "is there scheduling for content fragments",
        "Yes. When publishing a content fragment you can choose \"Now\" or \"Schedule\" and set an activation date. For jobs coming from Workday, publishing is done by the sync without author intervention.",
    ),
    (
        "reference should update if asset moved",
        "If a referenced asset (e.g. job image) is moved in the DAM, the reference should update. If the asset is deleted, the reference would break.",
    ),
    (
        "where are the filters for hot jobs listing page",
        "Filters are part of the Job Search component configuration. The component has filter labels and sort options (e.g. job categories, properties, job types, locations) that authors can edit in the component dialog.",
    ),
    (
        "do hiring events and job fairs automatically show and drop off",
        "Job postings from Workday automatically appear and disappear based on the sync; when a job is removed in Workday, the next sync removes it from AEM. Hiring events / job fairs and how they show or drop off depend on how those features are implemented (e.g. event content and scheduling); the Jobs KT focused on job postings and Workday sync.",
    ),
    (
        "select locations under the job search",
        "Select locations in the Job Search component are pulled from AEM: they are a combination of the country, state, and city fields from the job content fragments. New properties/locations appear automatically when new jobs (and thus new folder structure) come from Workday; authors do not add them manually.",
    ),
    (
        "new property, is the author responsible for adding",
        "No. When a new property and jobs are created in Workday, the sync creates the folder structure and content fragments in AEM. The author does not create or update locations/properties in the job search; they are driven by the content fragment data from Workday.",
    ),
]




def should_skip_answer(question: str) -> bool:
    q = (question or "").strip()
    if not q:
        return True
    if not q.endswith("?"):
        return True
    if q.lower().startswith("why "):
        return True
    return False


def find_answer(question: str) -> str | None:
    q = (question or "").lower().strip()
    for key, answer in ANSWER_MAP:
        if key in q:
            return answer
    return None


def main():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb["Research"]
    updated = 0
    for row_num in range(2, ws.max_row + 1):
        question = ws.cell(row=row_num, column=2).value
        if isinstance(question, str):
            question = question.strip()
        else:
            question = (question or "").strip()
        existing = ws.cell(row=row_num, column=5).value  # Answer column
        if existing and str(existing).strip():
            continue
        if should_skip_answer(question):
            continue
        answer = find_answer(question)
        if answer:
            ws.cell(row=row_num, column=5, value=answer)
            ws.cell(row=row_num, column=7, value="Adobe")  # Answered By
            updated += 1
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"Updated {updated} rows with Answer and Answered By = Adobe.")


if __name__ == "__main__":
    main()
