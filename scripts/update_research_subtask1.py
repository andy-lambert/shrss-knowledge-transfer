#!/usr/bin/env python3
"""
Subtask 1: Populate AI-Generated Question?, Confidence, and Reasoning in the Research sheet
based on transcript cross-reference and heuristics (technical terms, phrasing, length).
"""
import re
import csv
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
TRANSCRIPT_PATH = REPO_ROOT / "KT_Session_Transcripts" / "SHRSS_Adobe_KT_All_Session_Transcripts_Consolidated.md"
EXCEL_PATH = REPO_ROOT / "KT_Session_Follow_Up_Questions_RESEARCH_WORKING.xlsx"
EXPORT_CSV = REPO_ROOT / "KT_Research_export.csv"

SESSION_MARKERS = [
    ("Jobs", "## Session: Jobs — 2026-02-10", "## Session: Events — 2026-02-11"),
    ("Events", "## Session: Events — 2026-02-11", "## Session: Careers — 2026-02-12"),
    ("Careers", "## Session: Careers — 2026-02-12", "## Session: Tagging & Taxonomy — 2026-02-17"),
    ("Tagging_Taxonomy_Metadata_Gov", "## Session: Tagging & Taxonomy — 2026-02-17", "## Session: DAM — 2026-02-18"),
    ("DAM_Training_Usage_Admin", "## Session: DAM — 2026-02-18", "## Session: Shared Data — 2026-02-19"),
    ("Shared_Data", "## Session: Shared Data — 2026-02-19", "## Session: News — 2026-02-20"),
    ("News", "## Session: News — 2026-02-20", "## Session: Locations — 2026-02-23"),
    ("Locations", "## Session: Locations — 2026-02-23", "\x00"),
]

TECH_TERMS = (
    "sling", "osgi", "htl", "query builder", "search index", "jcr", "sling model",
    "content fragment model", "graphql", "dispatcher", "workflow", "metadata schema",
    "taxonomy", "cq:tags", "rendition", "renditions", "universal editor", "experience fragment",
    "content fragment", "cf model", "api data", "workday", "component dialog", "authoring",
    "dynamic components", "aria", "wcag", "ld json", "headless", "resolver", "servlet",
    "element id", "data layer", "focal point", "aspect ratio", "breakpoint", "responsive",
    "role-based", "production-ready", "mvp", "interim solution", "configurable", "statically configured",
    "relevance and ranking", "governance", "consolidat", "refactor", "generic list",
    "path/tag mapping", "acs commons", "namespace", "enumeration", "instrumented",
    "persisted query", "fragment model", "dplp", "dplt", "integration", "scheduler",
)

AI_PHRASES = [
    "functional differences between", "current intended", "validation steps should",
    "safest approach to", "how do we ensure", "how do we validate", "production-ready",
    "which components currently", "which components or services", "which fields are sourced",
    "does aem provide", "should .* be refactored", "what happens if a tag",
    "what is the roadmap", "are any .* incompatible", "what is the process to audit",
    "what is the difference in implementation", "what are the user roles",
    "how does tag selection affect", "how will property/location data",
    "what image renditions are generated", "what data is sent to the data layer",
    "is there a recommended", "is there documentation",
]


def norm(s):
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def words(s, min_len=3):
    return [w for w in re.findall(r"\b[a-z0-9]+\b", norm(s)) if len(w) >= min_len]


def phrase_matches(question, transcript, n=3):
    if not transcript:
        return 0
    qw = words(question)
    t_norm = norm(transcript)
    return sum(1 for i in range(len(qw) - n + 1) if " ".join(qw[i : i + n]) in t_norm)


def tech_count(question):
    ql = norm(question)
    return sum(1 for t in TECH_TERMS if t in ql)


def ai_phrase_score(question):
    ql = norm(question)
    return sum(1 for p in AI_PHRASES if p.replace(".*", " ") in ql or (".*" in p and re.search(p.replace(".*", ".*"), ql)))


def load_transcript_sections():
    text = TRANSCRIPT_PATH.read_text(encoding="utf-8")
    out = {}
    for name, start_marker, end_marker in SESSION_MARKERS:
        i = text.find(start_marker)
        j = text.find(end_marker, i + 1) if end_marker != "\x00" else len(text)
        out[name] = text[i:j] if i >= 0 and (j < 0 or j > i) else text[i:] if i >= 0 else ""
    return out


def compute_decision(question, session, transcript_by_session):
    if not question or not question.strip():
        return "FALSE", "0", "Empty question."
    transcript = transcript_by_session.get(session, "")
    pm = phrase_matches(question, transcript)
    tech = tech_count(question)
    aip = ai_phrase_score(question)
    wc = len(question.split())
    score = 0
    reasons = []
    if tech >= 2:
        score += 35
        reasons.append("multiple technical/AEM terms")
    elif tech == 1:
        score += 18
        reasons.append("technical term(s)")
    if aip >= 1:
        score += 28
        reasons.append("formal/AI-style phrasing")
    if wc >= 14:
        score += 12
        reasons.append("long, formal question")
    if pm >= 2:
        score += 22
        reasons.append("phrases match transcript (possible AI reword)")
    elif pm == 1:
        score += 8
        reasons.append("some transcript overlap")
    if "functional differences" in norm(question) or "current intended taxonomy" in norm(question):
        score += 18
        reasons.append("typical AI-generated taxonomy question")
    score = min(100, score)
    ai_generated = score >= 35
    confidence = str(score) if ai_generated else str(min(score + 8, 50))
    reasoning = "; ".join(reasons) if reasons else "No strong AI indicators; plausible author question."
    if ai_generated:
        reasoning = "Likely AI-generated: " + reasoning
    return str(ai_generated).upper(), confidence, reasoning


def main():
    import openpyxl
    transcript_by_session = load_transcript_sections()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb["Research"]
    # Row 1 = header; data rows 2..N
    decisions = []
    for row_num in range(2, ws.max_row + 1):
        session = (ws.cell(row=row_num, column=1).value or "")
        if isinstance(session, str):
            session = session.strip()
        else:
            session = str(session or "").strip()
        question = (ws.cell(row=row_num, column=2).value or "")
        if isinstance(question, str):
            question = question.strip()
        else:
            question = str(question or "").strip()
        decisions.append(compute_decision(question, session, transcript_by_session))
    true_count = sum(1 for d in decisions if d[0] == "TRUE")
    print(f"Computed {len(decisions)} decisions, AI-Generated TRUE: {true_count}")
    # Columns J=10, K=11, L=12 (1-based)
    for i, (ai_val, conf, reason) in enumerate(decisions):
        row_num = i + 2
        ws.cell(row=row_num, column=10, value=ai_val == "TRUE")
        ws.cell(row=row_num, column=11, value=conf)
        ws.cell(row=row_num, column=12, value=reason)
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"Updated {EXCEL_PATH}")


if __name__ == "__main__":
    main()
