#!/usr/bin/env python3
"""
Full Subtask 2 pass: Answer KT follow-up questions using ONLY:
- Implementation docs (01_STRUCTURAL_ARCHITECTURE, 02_CROSS_LAYER_INTERACTIONS, 04)
- Content/code references (paths from task)
- AEM canonical references and AdobeDocs-global-mapping (Experience League)
NO transcripts.

Skips: declarative sentences (no ?), questions starting with "Why".
Writes: Answer (col E), Answered By = Adobe (col G).
"""
from pathlib import Path
import re

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "KT_Session_Follow_Up_Questions_RESEARCH_WORKING.xlsx"

# Answer map: list of (session_or_None, question_substring_lower, answer_text)
# session_or_None = None means match any session; otherwise match that session only.
# First matching entry wins (order matters for overlap).
ANSWERS = [
    # ---- Jobs (from 01_STRUCTURAL, 02_CROSS_LAYER, 04) ----
    ("Jobs", "production-ready", "Per implementation analysis (01_STRUCTURAL_ARCHITECTURE, 04_IMPLEMENTATION_QUALITY), job-related components (jobsearch, job listing, etc.) are documented as part of the Search/Filter component set. The implementation has known issues: JobDetailsServlet has a GraphQL query syntax error (ISSUE-BACKEND-018); DeleteJobServlet and JobsCFUpdateServlet are missing authentication (P0). Production readiness should be confirmed with the development team and issue remediation status."),
    ("Jobs", "job-related components", "Job-related components in the SHRSS implementation include the Job Search component (jobsearch) and other job-dependent components. They are documented under Search & Filter Components in 01_STRUCTURAL_ARCHITECTURE.md (§2.5.5). Backend job Content Fragment management is documented in 02_CROSS_LAYER_INTERACTIONS.md (INT-BACKEND-002): servlets at /services/shrss/v1/jobs/getClientDetails, getJobIds, update, deleteJob."),
    ("Jobs", "configuration options role-based", "Implementation analysis does not document role-based configuration (author vs. admin) for job components. Template policies and component dialogs are under conf/shrss/settings/wcm/ (01_STRUCTURAL_ARCHITECTURE §4.2). Role-based behavior would be defined by AEM permissions and template policy configuration; specific SHRSS behavior would require code or config review."),
    ("Jobs", "filters driven dynamically from workday", "Per 02_CROSS_LAYER_INTERACTIONS (INT-UI-015), search and filter for Content Fragment lists (including Jobs) use client-side filtering with a backend servlet that constructs JCR or GraphQL queries. Filter options are derived from content (e.g. job category, properties, locations). Job data is populated via the job update servlet (INT-BACKEND-002); whether the source is “Workday” or another system is an integration detail. Filters are driven by the Content Fragments currently in AEM."),
    ("Jobs", "use job components on other sites", "Yes. Per 01_STRUCTURAL_ARCHITECTURE, jobsearch and related search/filter components are reusable. The CF Card List and search/filter pattern (INT-UI-015) support Events, News, Promotions, Jobs, Locations. Components can be placed on any page whose template allows them (template policies under conf/shrss/settings/wcm/policies)."),
    ("Jobs", "search page path", "Implementation docs do not specify a default Search Page path for the job search component. Component configuration (e.g. search page path) is typically stored in the component dialog and/or template policy. To confirm default or author-selectable path, review the jobsearch component dialog definition in the codebase."),
    ("Jobs", "load testing", "04_IMPLEMENTATION_QUALITY_ASSESSMENT documents unit test coverage (59.2% file coverage, models 93.8%, services 94.1%), 4 integration tests, and 8 Cypress UI tests. Load testing at scale is not documented in the implementation analysis. This would need to be confirmed with the project or QA team."),
    ("Jobs", "image cropping, scaling, focal", "AEM supports focal point and smart cropping for assets (see Experience League: Assets, Dynamic Media). Implementation analysis does not specify how job component images use focal point or scaling. Component image handling would be in the job card/listing HTL and Sling Model; for SHRSS-specific behavior, review the component code and any Adaptive Image Servlet or Dynamic Media usage."),
    ("Jobs", "search behave with no results", "Per 02_CROSS_LAYER_INTERACTIONS (INT-UI-015), when no results are found the backend returns an empty results array and the front end displays a “No results found” message. Search/filter uses AJAX and URL state management."),
    ("Jobs", "Job Listing Card.*incomplete or missing", "Implementation analysis does not document specific handling of incomplete or missing data in the Job Listing Card. General pattern: Sling Model exposes CF data to HTL; null/missing handling depends on the model and template. Review the job card component model and HTL for fallbacks."),
    ("Jobs", "search relevance and ranking", "INT-UI-015 describes sort and pagination (servlet returns results sorted by configured field, paginated). Relevance and ranking logic (e.g. keyword scoring) are not documented in the implementation analysis; they would depend on the backend servlet and query implementation."),
    ("Jobs", "layout responsive", "Standard AEM components and SHRSS UI use clientlibs and responsive patterns. 01_STRUCTURAL_ARCHITECTURE references responsive behavior; specific breakpoints for job components are not detailed. See Front-End Development and Clientlibs in AEM_CANONICAL_REFERENCES.md."),
    ("Jobs", "user roles created for this website", "RepoInit and service user mappings are documented in 01_STRUCTURAL_ARCHITECTURE (§4.1). Author-facing roles (e.g. author, admin) are typically defined in RepoInit (config.author/config.publish RepositoryInitializer). The implementation analysis does not list specific “website” user roles; these would be in the repoinit configuration."),
    ("Jobs", "data layer on job click", "Implementation analysis does not document analytics data layer or tracking events for job card clicks. This would be implemented in client-side JavaScript or in the component’s clientlib. For SHRSS-specific behavior, review the job component clientlib and any Tealium/analytics integration."),
    ("Jobs", "category/location/property/type has no jobs", "Per INT-UI-015, filter options are built from available content. If a category/location/property/type has no jobs, it may not appear as a filter option or would show zero results; exact behavior depends on the component and servlet implementation."),
    ("Jobs", "image renditions", "AEM Assets generate OOTB renditions; custom renditions can be defined. Implementation analysis does not specify which renditions are used by job components or whether aspect ratios are enforced. See AEM Assets and Dynamic Media in AEM_CANONICAL_REFERENCES.md; SHRSS-specific rendition usage would be in component code or image configuration."),
    ("Jobs", "element ID or tracking event for job card", "Not documented in implementation analysis. Element IDs and tracking events are typically set in HTL and clientlibs; review the job card component for data attributes or IDs used for analytics."),
    ("Jobs", "DAM folder structure for job-related images", "01_STRUCTURAL_ARCHITECTURE does not specify a finalized DAM folder structure for job images. Content Fragment models and templates are under conf/shrss/settings/. Job CFs are stored under the content fragment hierarchy; image configuration may be in component dialogs or a dedicated config. For current structure, review ui.content or content package."),
    ("Jobs", "create a page.*inherit header/footer", "Page templates define structure; header/footer are typically provided by the page template or experience fragments. 01_STRUCTURAL_ARCHITECTURE (§4.2) references templates under conf/shrss/settings/wcm/templates. Inheritance of header/footer is template-dependent; review the Careers/job page template structure."),
    ("Jobs", "access to configuration pages for job", "Access is governed by AEM permissions and template policy. Implementation analysis does not enumerate “configuration pages” for jobs. Who can configure job components is determined by CUG/permissions and policy assignment; see AEM Security and Touch UI in AEM_CANONICAL_REFERENCES.md."),
    ("Jobs", "filters interdependent", "Not documented. Filter interdependence (e.g. cascading) would be in the search/filter clientlib and backend servlet (INT-UI-015)."),
    ("Jobs", "Is API Data.*unchecked", "Implementation docs describe job CF create/update via servlets (INT-BACKEND-002). “Is API data” is a Content Fragment model field; when unchecked, the CF is typically treated as author-created. Sync/update behavior for such CFs depends on the integration (e.g. Workday sync) and is not detailed in the cross-layer doc."),
    ("Jobs", "address from API.*where.*used", "Job Content Fragment management (INT-BACKEND-002) maps JSON to CF properties. Address or location fields from the API would be stored on the CF and used wherever the component or servlet reads those properties (e.g. display, folder structure, filters). Exact usage is in the job component model and servlet logic."),
    ("Jobs", "job detail page URL exists but job no longer", "When a job is deleted via DeleteJobServlet (INT-BACKEND-002), the Content Fragment is removed from JCR. If a job detail page URL exists but the job CF no longer exists, the page or component would typically 404 or show empty; the doc does not describe a custom 404 for expired jobs."),
    ("Jobs", "templates.*returning applicants, team members", "Implementation analysis does not document templates for “returning applicants” or “team members.” These would be custom or out-of-scope for the documented job CF and servlet flows."),
    ("Jobs", "filters for hot jobs listing", "Filters are part of the Job Search component configuration (INT-UI-015; Search & Filter Components in 01_STRUCTURAL_ARCHITECTURE). Filter labels and options are configurable in the component dialog; the “hot jobs listing” page would use the same or a related component with its own filter config."),
    ("Jobs", "Job Zip Code used", "Not documented in implementation analysis. Zip code would be a field on the Job Content Fragment model if present; usage would be in the component or query logic. Check the Job CF model and job-related components in the codebase."),
    ("Jobs", "tags.*internal metadata.*query content", "In AEM, tags (cq:tags) can be used to query content; Query Builder and JCR queries support tag predicates. Implementation analysis does not specify whether SHRSS job or list components query by tags; see Administering Tags and Query Builder in AEM_CANONICAL_REFERENCES.md."),
    # ---- Events ----
    ("Events", "fields.*card and detail page.*managed by status", "Implementation analysis does not enumerate which Event card/detail fields are status-driven. Event CF model and components would define this; see staging/ui/STRUCTURAL_UI.md and Event Content Fragment usage in 01_STRUCTURAL_ARCHITECTURE (§2.5.5, §2.6)."),
    ("Events", "schedule events to automatically publish and unpublish", "AEM supports scheduled activation/deactivation for pages and content. Content Fragments can be scheduled per AEM capabilities. See Content Fragments and Replication in AEM_CANONICAL_REFERENCES.md. SHRSS-specific event scheduling would be in workflow or event CF model and component logic."),
    ("Events", "events automatically drop off", "Automatic drop-off after end date depends on implementation: either scheduled unpublish or a query that filters by date. 02_CROSS_LAYER documents search/filter and CF listing (INT-UI-015); date-based filtering would be in the servlet query. Exact behavior is implementation-specific."),
    ("Events", "property-level timezones", "Timezone display is typically handled in the component or front end (venue timezone vs. browser). Implementation analysis does not document property-level timezone configuration; this would require code/config review."),
    ("Events", "query.*events.*other components", "INT-UI-015 describes dynamic queries for CF lists (Events, News, Jobs, etc.) with backend servlet and AJAX. Events can be queried and displayed in other components (carousel, grid) if those components use the same or similar servlet/query pattern; see CF Card List and search/filter components."),
    ("Events", "version history.*Content Fragment", "AEM Content Fragments support versioning. Implementation analysis does not specify SHRSS workflow for CF version history or rollback. See Content Fragments in AEM_CANONICAL_REFERENCES.md."),
    ("Events", "Event Detail.*SEO.*schema", "Schema markup (Event, Venue, Offer) is typically implemented in the component HTL or via a dedicated schema service. Implementation analysis does not document event-specific schema; see structured data and SEO in Experience League."),
    ("Events", "Content Fragment model.*Event", "01_STRUCTURAL_ARCHITECTURE (§2.6) lists Event Content Fragment Model as one of six CF models (conf/shrss/settings/dam/cfm/models). GraphQL exposes CFs for headless consumption. Event-specific fields and behavior are in the model and consuming components."),
    # ---- Careers ----
    ("Careers", "accessibility controls.*captions.*ARIA", "Accessibility (captions, ARIA) is a component-level and front-end concern. Implementation analysis does not document which Careers components support which accessibility features. See WCAG and component implementation in the codebase and AEM Core Components."),
    ("Careers", "CF Model is modified.*what happens to existing", "When a Content Fragment Model is changed (field added/removed), existing fragments retain stored content. New fields appear empty; removed fields may no longer show in the editor or API. See Content Fragment Model docs in AEM_CANONICAL_REFERENCES.md; test in lower environment before production changes."),
    ("Careers", "responsive renditions.*Dynamic Media.*Adaptive", "01_STRUCTURAL_ARCHITECTURE does not specify whether SHRSS uses Dynamic Media, Adaptive Image Servlet, or static crops for Careers/promotion components. See Assets and Dynamic Media in AEM_CANONICAL_REFERENCES.md; SHRSS-specific behavior is in component and asset configuration."),
    ("Careers", "DPLT.*update.*cascade", "DPLT (location/property data) integration is not fully detailed in the cross-layer doc. If DPLT is the source for location CFs or shared data, updates would cascade only if the implementation syncs or references that source; see Locations and Shared Data sections and integration code."),
    ("Careers", "Promotion.*Content Fragment", "01_STRUCTURAL_ARCHITECTURE (§2.6) lists Promotion Content Fragment Model. Promotions are rendered via cfcard, cfcardlist, and dedicated CF components. Query and filter behavior for Promotions follows INT-UI-015 (search/filter, AJAX, backend servlet)."),
    # ---- Tagging ----
    ("Tagging_Taxonomy_Metadata_Gov", "taxonomy architecture.*SHRSS namespace", "Tags in AEM live under /content/cq:tags; SHRSS-specific branches (e.g. Categories, Event Categories) are implementation-defined. 01_STRUCTURAL_ARCHITECTURE does not enumerate tag namespaces. To see current taxonomy: Tools → General → Tagging and review conf/shrss or content where tag paths are configured."),
    ("Tagging_Taxonomy_Metadata_Gov", "functional differences between.*Categories.*Category", "From an AEM perspective, all are cq:Tag nodes; behavioral differences come from how the implementation uses them (e.g. which components and queries reference which paths). See Administering Tags in AEM_CANONICAL_REFERENCES.md."),
    ("Tagging_Taxonomy_Metadata_Gov", "safest approach to consolidating.*tag", "Recommended: (1) Inventory usage via Tag console References and codebase search for tag paths. (2) Prefer move/merge over delete so cq:tags references update. (3) Update custom properties explicitly if used. (4) Test in lower environment. See AEM tagging and taxonomy best practices in Experience League."),
    ("Tagging_Taxonomy_Metadata_Gov", "tag in use is deleted", "Content nodes keep the old tag ID; the ID no longer resolves to a valid tag. Tag pickers may show missing values; tag-based queries may not return expected items. Prefer move/merge over delete. See AEM_CANONICAL_REFERENCES and tagging docs."),
    ("Tagging_Taxonomy_Metadata_Gov", "AEM provide dependency warnings before tag deletion", "The Tag console shows References for a tag. AEM may warn on delete when references exist; it does not automatically update custom properties. Governance relies on permissions and process. See Administering Tags."),
    ("Tagging_Taxonomy_Metadata_Gov", "Universal Editor", "Universal Editor support depends on component instrumentation (see Universal Editor in AEM_CANONICAL_REFERENCES.md). Implementation analysis does not state which SHRSS functionality is unsupported in Universal Editor; this requires reviewing components against Universal Editor guidelines."),
    ("Tagging_Taxonomy_Metadata_Gov", "tag selection in a Content Fragment affect dynamic listings", "If list components or GraphQL queries filter by cq:tags (or a tag-backed property), changing tags on a CF changes whether it appears in those listings. Implementation-specific; see INT-UI-015 and component query logic."),
    ("Tagging_Taxonomy_Metadata_Gov", "metadata schema.*tag", "Asset metadata schemas can define tag-driven fields (Tools → Assets → Metadata Schemas). Which SHRSS metadata fields are tag-driven is in conf/shrss/settings/dam (01_STRUCTURAL_ARCHITECTURE §4.2)."),
    # ---- DAM ----
    ("DAM_Training_Usage_Admin", "metadata fields.*mandatory upon upload", "Mandatory metadata is defined in metadata schemas (conf/shrss/settings/dam). Implementation analysis does not list which fields are mandatory; review metadata schema configuration in the project. See Metadata Schemas in AEM_CANONICAL_REFERENCES.md."),
    ("DAM_Training_Usage_Admin", "static renditions.*chosen.*components", "Components reference assets and optionally a specific rendition or rely on OOTB image component behavior (e.g. responsive image, Dynamic Media). Which renditions are “chosen” depends on the component (HTL and Sling Model). See Assets API and clientlibs in AEM_CANONICAL_REFERENCES.md."),
    ("DAM_Training_Usage_Admin", "recommended image rendition sizes", "Implementation analysis does not specify recommended rendition sizes for SHRSS (cards, banners, hero, etc.). This is typically defined in design specs and component configuration; see Dynamic Media and Assets in Experience League."),
    ("DAM_Training_Usage_Admin", "Dynamic Media.*static renditions", "AEM supports both static renditions and Dynamic Media. Transition plan (static → Dynamic Media) is project-specific and not documented in implementation analysis. See Dynamic Media in AEM_CANONICAL_REFERENCES.md."),
    # ---- Shared Data ----
    ("Shared_Data", "decouple the footer", "Header/footer are typically provided by page template or Experience Fragments. “Decoupling” the footer would mean making it configurable or fragment-based so not every page is tied to one footer. Implementation analysis references XF and shared components; exact footer structure is in template and XF usage."),
    ("Shared_Data", "validate which components are allowed on which templates", "Template policies (conf/shrss/settings/wcm/policies) define which components are allowed on a template. Authors see only allowed components in the parsys. See Templates and Overlays in AEM_CANONICAL_REFERENCES.md."),
    ("Shared_Data", "enable/disable components at the template level", "Yes. Template policies control which components are allowed in each container. Changing policy enables/disables components without creating new templates. See 01_STRUCTURAL_ARCHITECTURE §4.2 and AEM template documentation."),
    ("Shared_Data", "CF references break if CFs are moved", "AEM can update references when content is moved, depending on reference type and implementation. Implementation analysis does not specify SHRSS behavior for CF moves; test in lower environment. See Content Fragments and reference handling in Experience League."),
    ("Shared_Data", "component audit.*clientlibs.*Sling", "01_STRUCTURAL_ARCHITECTURE and staging/ui/STRUCTURAL_UI.md provide a component inventory with patterns (backend servlet, clientlib, HTL, Sling Model). A full audit showing which components rely on which technical layers is in the structural docs; see §2.5 and staging UI docs."),
    ("Shared_Data", "navigation.*cached and invalidated", "Dispatcher and CDN caching are documented in 01_STRUCTURAL_ARCHITECTURE (§3). Navigation (e.g. from XF or shared content) is invalidated when that content is published; exact invalidation rules are in dispatcher cache config. See Dispatcher in AEM_CANONICAL_REFERENCES.md."),
    # ---- Locations ----
    ("Locations", "fields are sourced from DPLT", "Implementation analysis references location Content Fragments and Location Data Export (INT-BACKEND-006). Which Location CF fields are sourced from DPLT vs. author-editable is implementation-specific; see Location CF model and any DPLT integration code."),
    ("Locations", "location.*status.*visibility", "If location status is stored on the CF or in DPLT, visibility/behavior in AEM would be determined by component or query logic that filters by status. Not explicitly documented in cross-layer doc; review location components and queries."),
    ("Locations", "location added to DPLT.*automatically appear", "Automatic appearance of new DPLT locations in AEM depends on the integration (sync job or API). INT-BACKEND-006 describes location data export, not DPLT import; DPLT-to-AEM flow would be in a separate integration. Confirm with integration documentation or code."),
    ("Locations", "Location Content Fragment.*read-only vs", "Which Location CF fields are read-only vs. author-editable is defined in the Content Fragment Model and any logic that populates from DPLT. See conf/shrss/settings/dam/cfm/models and location integration."),
    ("Locations", "Content Fragment Model.*update.*process", "Adding a new field to a CF model requires: (1) Update the model in conf. (2) Deploy. (3) Update any Sling Models, HTL, or GraphQL queries that need to expose/use the field. (4) Optionally backfill existing fragments. See Content Fragment Model docs in AEM_CANONICAL_REFERENCES.md. Who can modify models is governed by AEM permissions."),
    ("Locations", "Google Map.*Destination Search", "Implementation analysis lists googlemap and destinationsearch as separate components (§2.5.5). Functional differences (e.g. manual country selection) would be in component dialogs and implementation; review both components in the codebase."),
    ("Locations", "Booking Widget.*Experience Fragment", "Integration components (e.g. booking widget) may be embedded in pages or XFs. Whether the Booking Widget is configured via XF is implementation-specific; see 01_STRUCTURAL_ARCHITECTURE §2.5.4 and XF usage."),
    # More Jobs
    ("Jobs", "default search behaviors configurable", "INT-UI-015: sort options (e.g. most recent, job title, location) are configured on the Job Search component. Whether location-first or keyword-first is configurable depends on component dialog and servlet; see 02_CROSS_LAYER_INTERACTIONS."),
    ("Jobs", "page template.*support additional content", "Templates and policies under conf/shrss/settings/wcm define allowed components. Whether a job template supports additional blocks is template-specific; see 01_STRUCTURAL_ARCHITECTURE §4.2."),
    ("Jobs", "custom 404 for expired jobs", "Implementation analysis does not document a custom 404 for expired jobs. When a job CF is deleted, 404 behavior depends on URL and error handler config."),
    ("Jobs", "version history for job overrides", "AEM Content Fragments support versioning; author overrides on the CF are versioned with the CF. See Content Fragments in AEM_CANONICAL_REFERENCES.md."),
    ("Jobs", "Hiring Event detail page template", "Careers/job page templates are under conf/shrss/settings/wcm/templates. A dedicated Hiring Event detail template is not specified in implementation analysis; review template list in content or codebase."),
    ("Jobs", "What does type do", "Job type (e.g. full-time) is a field on the Job CF; it affects filters and possibly display. See INT-UI-015 and job component model."),
    # More Events
    ("Events", "dynamically query and display events", "INT-UI-015: Events are queried via backend servlet; results can be shown in CF Card List, carousel, or grid if components use the same servlet/query."),
    ("Events", "event listings using queries", "Search/filter supports query parameters (tags, metadata, path). Listings by tags/metadata require servlet query logic; see INT-UI-015 and staging backend."),
    ("Events", "filter state persist across navigation", "INT-UI-015: URL state (pushState) reflects filters; back/forward triggers popstate and re-fetch. Cross-navigation persistence depends on URL params."),
    ("Events", "filtered results URL-driven", "Yes. INT-UI-015 describes query parameters in the URL for deep linking and sharable filtered views."),
    ("Events", "pagination vs.*Load More", "INT-UI-015 documents pagination (configurable page size). Load More or infinite scroll would be a front-end variant; implementation-specific."),
    ("Events", "version history.*Content Fragment", "AEM Content Fragments support versioning and rollback. See Content Fragments in AEM_CANONICAL_REFERENCES.md."),
    ("Events", "Content Fragment.*hybrid.*Experience Fragment", "01_STRUCTURAL_ARCHITECTURE lists Event as a CF model; XFs are used for shared layout. Architecture choice; see Content Fragments and Experience Fragments in AEM_CANONICAL_REFERENCES.md."),
    ("Events", "AEM automatically generate structured data for Event", "AEM does not auto-generate Event schema; implement in component or schema service. See structured data in Experience League."),
    ("Events", "master event taxonomy", "Taxonomy is under /content/cq:tags and in Event CF model/dialogs. See 01_STRUCTURAL_ARCHITECTURE §4.2 and Tagging."),
    # More Tagging
    ("Tagging_Taxonomy_Metadata_Gov", "components or services.*reference", "Implementation analysis does not list component–tag mapping. Search codebase for tag paths and cq:tags; use Tag console References."),
    ("Tagging_Taxonomy_Metadata_Gov", "validation steps.*restructuring", "Before restructuring: inventory usage (Tag References, code search); test search/navigation in lower env; prefer move/merge. See AEM tagging best practices."),
    ("Tagging_Taxonomy_Metadata_Gov", "Generic Lists", "Generic Lists (e.g. ACS Commons) are typically under /conf or /etc. See 01_STRUCTURAL_ARCHITECTURE §4.2 and codebase for list references."),
    ("Tagging_Taxonomy_Metadata_Gov", "Category fields.*refactored to reference", "Using centralized tags in CF models is a best practice. Refactoring requires model change, content migration, and component/query updates. See AEM tagging and CF model docs."),
    ("Tagging_Taxonomy_Metadata_Gov", "difference.*CQ:tags.*Category", "cq:tags is standard AEM tag property; Category dropdown can be enum or tag-backed; metadata schema tag fields store tag IDs. See Administering Tags and Metadata Schemas."),
    ("Tagging_Taxonomy_Metadata_Gov", "Category a required field", "Required fields are per CF model. See conf/shrss/settings/dam/cfm/models."),
    ("Tagging_Taxonomy_Metadata_Gov", "audit tag usage", "Tag console References show usage. Asset metadata bulk export to CSV is in AEM Assets; see Experience League."),
    ("Tagging_Taxonomy_Metadata_Gov", "naming convention for tag", "Adobe does not mandate one; best practice: meaningful titles, stable IDs, hierarchy. See Administering Tags."),
    ("Tagging_Taxonomy_Metadata_Gov", "validate.*dynamic components break", "Regression test all tag-using features after consolidation; run in lower env. No automatic guarantee."),
    # More DAM
    ("DAM_Training_Usage_Admin", "metadata.*language", "Language can be in metadata or folder structure. Metadata-based language (EN/ES/FR) is supported if workflows use it. See AEM Assets metadata and translation."),
    ("DAM_Training_Usage_Admin", "prevent.*old or duplicate assets", "Governance: naming, folders, metadata. Component references (path/UUID) and move behavior are in Assets docs. See AEM_CANONICAL_REFERENCES.md."),
    ("DAM_Training_Usage_Admin", "correct rendition chosen", "Image components use default or responsive config. Correct rendition is component/config dependent; see OOTB image and Dynamic Media docs."),
    # More Shared_Data
    ("Shared_Data", "force authors into structured content", "Template policy can restrict to CF-based components and limit free-form content. See Templates and AEM_CANONICAL_REFERENCES.md."),
    ("Shared_Data", "components support personalization", "AEM supports Context Hub and Target. Which SHRSS components use it is implementation-specific; see 01_STRUCTURAL_ARCHITECTURE."),
    ("Shared_Data", "restrict.*components on.*templates", "Yes. Template policies define allowed components per container. See 01_STRUCTURAL_ARCHITECTURE §4.2 and Templates."),
    ("Shared_Data", "policy inheritance", "Policies can inherit; override is via Sling Resource Merger. See Overlays and AEM_CANONICAL_REFERENCES.md."),
    ("Shared_Data", "caching and invalidation when CFs", "Dispatcher/CDN invalidate on publish; CF update and publish may require page invalidation. See 01_STRUCTURAL_ARCHITECTURE §3 and Replication/CDN."),
    # More Locations
    ("Locations", "steps to add locations", "Add location component to page; configure path/filter in dialog. See 01_STRUCTURAL_ARCHITECTURE §2.5.4 (googlemap, destinationsearch)."),
    ("Locations", "add locations without.*accordions", "Component dialog may offer display modes (e.g. map only, list). Implementation-specific."),
    ("Locations", "location not yet in the DPLT", "If sourced from DPLT, add to DPLT first or create Location CF manually if allowed. See DPLT integration."),
    ("Locations", "criteria.*display specific locations", "Filtering (LOB, geo, etc.) is in component and backend query. Configure in dialog or query; see INT-UI-015."),
    ("Locations", "permission to modify Content Fragment Models", "CF models are in config (code); change via Cloud Manager. Permissions are typically developer/admin. See 01_STRUCTURAL_ARCHITECTURE §4.2."),
    ("Locations", "new field.*model.*front-end", "Add field to model, deploy, update Sling Model/HTL and GraphQL if needed, optionally backfill. See Content Fragment and component docs."),
]

# Normalize for matching: lowercase, collapse whitespace
def norm(s):
    return re.sub(r"\s+", " ", (s or "").lower().strip())

def should_skip(question: str) -> bool:
    q = norm(question)
    if not q or not q.endswith("?"):
        return True
    if q.startswith("why "):
        return True
    return False

def find_answer(session: str, question: str) -> str | None:
    q = norm(question)
    s = (session or "").strip()
    for sess, sub, answer in ANSWERS:
        if sub not in q:
            continue
        if sess is not None and s != sess:
            continue
        return answer
    return None

def main():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb["Research"]
    updated = 0
    skipped = 0
    for row in range(2, ws.max_row + 1):
        session = (ws.cell(row=row, column=1).value or "")
        session = str(session).strip() if session else ""
        question = (ws.cell(row=row, column=2).value or "")
        question = str(question).strip() if question else ""
        if should_skip(question):
            skipped += 1
            continue
        answer = find_answer(session, question)
        if answer:
            ws.cell(row=row, column=5, value=answer)
            ws.cell(row=row, column=7, value="Adobe")
            updated += 1
        else:
            # Clear any prior answer (e.g. transcript-sourced) when we have no doc-sourced answer
            ws.cell(row=row, column=5, value=None)
            ws.cell(row=row, column=7, value=None)
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"Subtask 2 full pass complete. Answered: {updated}. Skipped (declarative/Why): {skipped}.")

if __name__ == "__main__":
    main()
