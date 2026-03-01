#!/usr/bin/env python3
"""Combine all session sheets into one 'All_Sessions' sheet with a Session column (source sheet name)."""
from pathlib import Path
import openpyxl

WB_PATH = Path(__file__).resolve().parent / "KT_Session_Follow_Up_Questions_Analysis_20260226.xlsx"
OUTPUT_HEADERS = ["Session", "Question", "Date Asked", "Asked By", "Answer", "Answered On", "Answered By", "Status"]


def normalize_header(h):
    if h is None:
        return ""
    return str(h).strip()


def main():
    wb = openpyxl.load_workbook(WB_PATH)
    # Skip if All_Sessions already exists and remove it so we recreate
    if "All_Sessions" in wb.sheetnames:
        del wb["All_Sessions"]
    ws_all = wb.create_sheet("All_Sessions", 0)
    for col, h in enumerate(OUTPUT_HEADERS, 1):
        ws_all.cell(row=1, column=col, value=h)
    next_row = 2
    for sheet_name in wb.sheetnames:
        if sheet_name == "All_Sessions":
            continue
        ws = wb[sheet_name]
        headers = [normalize_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        # Map header name -> column index (1-based)
        col_map = {}
        for c, h in enumerate(headers, 1):
            col_map[h] = c
        for r in range(2, ws.max_row + 1):
            row_vals = {}
            for h, c in col_map.items():
                row_vals[h] = ws.cell(row=r, column=c).value
            # Build output row: Session first, then standard order (match headers with strip)
            out = [sheet_name]
            for out_h in OUTPUT_HEADERS[1:]:  # skip Session
                val = None
                out_h_norm = out_h.strip().lower()
                for k, v in row_vals.items():
                    if k.strip().lower() == out_h_norm:
                        val = v
                        break
                out.append(val)
            for col, val in enumerate(out, 1):
                ws_all.cell(row=next_row, column=col, value=val)
            next_row += 1
    wb.save(WB_PATH)
    print(f"Combined {next_row - 2} rows into sheet 'All_Sessions'. Headers: {OUTPUT_HEADERS}")


if __name__ == "__main__":
    main()
