#!/usr/bin/env python3
"""
Find 'From Transcript' rows in session sheets, search consolidated transcript for
the question, extract the answer (with timestamp prefix), update Answer column in
session sheet and in All_Questions.
"""
import re
from pathlib import Path

EXCEL_PATH = Path("/Users/lambert/Documents/Projects/SHRSS/SHRSS_Knowledge_Transfer/KT_Session_Question_Review/SHRSS_Adobe_KT_Session_Questions.xlsx")
TRANSCRIPT_PATH = Path("/Users/lambert/Documents/Projects/SHRSS/SHRSS_Knowledge_Transfer/KT_Session_Transcripts/SHRSS_Adobe_KT_All_Session_Transcripts_Consolidated.md")

# Sheet name -> transcript section header start (unique match)
SHEET_TO_SESSION = {
    "Jobs": "## Session: Jobs —",
    "Events": "## Session: Events —",
    "Careers": "## Session: Careers —",
    "Tagging_Taxonomy_Metadata_Gov": "## Session: Tagging & Taxonomy —",
    "DAM_Training_Usage_Guide_Admins": "## Session: DAM —",
    "Shared_Data": "## Session: Shared Data —",
    "News": "## Session: News —",
    "Locations": "## Session: Locations —",
    "Page_Templates": "## Session: Page Templates —",
    "Media": "## Session: Media —",
    "Navigation_and_Data_Display": "## Session: Nav and Data Displays —",
    "LOB_Specific": "## Session: LOB Specific Components —",
    "Additional_Components": "## Session: Additional Components —",
}

# Column indices (1-based) for session sheets: Date Asked, Status, Flag, Question, Answer, Asked By, Answered By, Answered On
COL_QUESTION = 4
COL_ANSWER = 5
COL_ASKED_BY = 6

MAX_CELL_CHARS = 32000  # Excel limit


def normalize(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    # Normalize apostrophes for transcript match (Excel may have ', transcript has \'
    s = s.replace("\u2019", "'").replace("'", "'")
    return s


def parse_transcript_sections(path: Path) -> dict:
    """Return dict: session_header_start -> full section text (between this header and next ## Session)."""
    text = path.read_text(encoding="utf-8", errors="replace")
    sections = {}
    current_key = None
    current_lines = []
    for line in text.splitlines():
        if line.startswith("## Session:"):
            if current_key is not None:
                sections[current_key] = "\n".join(current_lines)
            # Find which key this matches
            current_key = None
            for k, prefix in SHEET_TO_SESSION.items():
                if line.startswith(prefix):
                    current_key = k
                    break
            # If no key (e.g. "Locations, Day 2"), keep previous section or skip
            if current_key is None:
                current_key = "_other"
            current_lines = [line]
        else:
            current_lines.append(line)
    if current_key is not None:
        sections[current_key] = "\n".join(current_lines)
    return sections


def find_question_position(section_text: str, question: str) -> int:
    """Return character offset where question (or a substantial substring) appears, or -1."""
    q = normalize(question)
    if not q:
        return -1
    section_norm = section_text.replace("\\'", "'").replace("\u2019", "'")
    section_lower = section_norm.lower()
    q_lower = q.lower()
    words = q.split()
    # Prefer distinctive mid-question phrases (sliding window) to avoid matching generic starts like "Can you explain what"
    for w in (5, 4):
        for i in range(len(words) - w + 1):
            phrase = " ".join(words[i : i + w])
            if len(phrase) >= 18 and phrase.lower() in section_lower:
                return section_lower.index(phrase.lower())
    # Try full question then progressively shorter prefix
    for length in (len(q), 100, 80, 60, 50, 40, 35, 30):
        if length < 12:
            break
        sub = q[:length].strip()
        if sub and sub.lower() in section_lower:
            return section_lower.index(sub.lower())
    # First N words (at least 6 so we don't match generic "Can you explain what")
    for n in (10, 8, 7, 6):
        if len(words) >= n:
            phrase = " ".join(words[:n])
            if len(phrase) >= 20 and phrase.lower() in section_lower:
                return section_lower.index(phrase.lower())
    # Shorter prefix only if at least 25 chars to reduce false matches
    for length in (25, 20):
        if len(q) >= length:
            sub = q[:length].strip()
            if sub and sub.lower() in section_lower:
                return section_lower.index(sub.lower())
    return -1


# Speaker line: **Name** M:SS [optional dialogue on same line]. Match from start so we detect speaker even when dialogue follows.
SPEAKER_RE = re.compile(r"^(\*\*[^*]+\*\*|\S.+?\*\*)\s+(\d{1,2}:\d{2}(?::\d{2})?)\s*(.*)$")


def find_next_speaker_line_after(lines: list, after_char_offset: int) -> tuple:
    """
    Given lines (with \n implied), find the first speaker line that starts at or after after_char_offset.
    Return (line_index, match) or (-1, None). match.group(1) is name part, match.group(2) is timestamp.
    """
    offset = 0
    for i, line in enumerate(lines):
        line_len = len(line) + 1  # +1 for newline
        if offset + len(line) >= after_char_offset:
            m = SPEAKER_RE.match(line.strip())
            if m:
                return (i, m)
        offset += line_len
    return (-1, None)


def extract_answer_block(lines: list, start_line_idx: int, speaker_match) -> str:
    """
    From start_line_idx we have a speaker line (e.g. **Daniela Tea** 12:34 or **Name** 12:34 dialogue).
    Collect that line's dialogue (group(3) if any) and all following lines until the next speaker line.
    Return string: "**Name** timestamp dialogue..."
    """
    timestamp = speaker_match.group(2)
    name_part = speaker_match.group(1).strip()
    prefix = f"{name_part} {timestamp}"
    rest_of_line = (speaker_match.group(3).strip() if speaker_match.lastindex >= 3 else "") or ""
    parts = [rest_of_line] if rest_of_line else []
    for i in range(start_line_idx + 1, len(lines)):
        line = lines[i]
        if SPEAKER_RE.match(line.strip()):
            break
        parts.append(line)
    dialogue = "\n".join(parts).strip()
    return f"{prefix} {dialogue}".strip() if dialogue else prefix


def _block_contains_question(lines: list, speaker_line_idx: int, question: str) -> bool:
    """Return True if the speaker block starting at speaker_line_idx contains the question text."""
    q_norm = normalize(question).lower()
    if not q_norm or len(q_norm) < 15:
        return False
    # Collect this speaker's block (this line + following until next speaker)
    chunk_parts = [lines[speaker_line_idx]]
    for i in range(speaker_line_idx + 1, len(lines)):
        if SPEAKER_RE.match(lines[i].strip()):
            break
        chunk_parts.append(lines[i])
    chunk = " ".join(chunk_parts).replace("\\'", "'").lower()
    # Question is in this block if a substantial substring (e.g. 20+ chars) appears
    for length in (len(q_norm), 40, 30, 20):
        if length < 20:
            break
        sub = q_norm[:length]
        if sub in chunk:
            return True
    return False


def find_answer_in_section(section_text: str, question: str) -> str | None:
    """
    Find question in section, then the next speaker's response. Return prefixed answer or None.
    If the question appears inside a speaker's block (e.g. asker's line), return the *following* speaker's block.
    """
    pos = find_question_position(section_text, question)
    if pos < 0:
        return None
    lines = section_text.splitlines()
    # Find line that contains position pos (offset is start of line i)
    offset = 0
    start_line_idx = 0
    for i, line in enumerate(lines):
        line_end = offset + len(line)
        if offset <= pos < line_end:
            start_line_idx = i
            break
        offset = line_end + 1  # +1 for newline
    # Find first speaker line at or after start_line_idx
    first_speaker_idx = -1
    first_match = None
    for i in range(start_line_idx, len(lines)):
        m = SPEAKER_RE.match(lines[i].strip())
        if m:
            first_speaker_idx = i
            first_match = m
            break
    if first_speaker_idx < 0:
        return None
    # If the question is in this speaker's block, they're the asker — use the next speaker's block (the answer).
    if _block_contains_question(lines, first_speaker_idx, question):
        for i in range(first_speaker_idx + 1, len(lines)):
            m = SPEAKER_RE.match(lines[i].strip())
            if m:
                return extract_answer_block(lines, i, m)
        return None
    return extract_answer_block(lines, first_speaker_idx, first_match)


def main():
    import openpyxl

    sections = parse_transcript_sections(TRANSCRIPT_PATH)
    # Search full transcript (question may have been asked in any session)
    full_transcript = "\n".join(sections.get(k, "") for k in SHEET_TO_SESSION)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=False)

    session_sheets = [n for n in wb.sheetnames if n not in ("Highlighted_Questions", "All_Questions")]
    all_ws = wb["All_Questions"]
    all_session_col = 1
    all_question_col = 5
    all_answer_col = 6

    all_question_to_row = {}  # (session, norm_question) -> list of row indices in All_Questions
    for r in range(2, all_ws.max_row + 1):
        sess = all_ws.cell(row=r, column=all_session_col).value
        q = all_ws.cell(row=r, column=all_question_col).value
        key = (str(sess).strip() if sess else "", normalize(q) if q else "")
        all_question_to_row.setdefault(key, []).append(r)

    filled = 0
    not_found = 0

    for sheet_name in session_sheets:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            asked_by = ws.cell(row=r, column=COL_ASKED_BY).value
            if not asked_by or str(asked_by).strip() != "From Transcript":
                continue
            question = ws.cell(row=r, column=COL_QUESTION).value
            if not question:
                continue
            answer = find_answer_in_section(full_transcript, str(question))
            if not answer:
                not_found += 1
                continue
            if len(answer) > MAX_CELL_CHARS:
                answer = answer[: MAX_CELL_CHARS - 3] + "..."
            ws.cell(row=r, column=COL_ANSWER, value=answer)
            filled += 1
            # Update All_Questions
            q_norm = normalize(question)
            key = (sheet_name, q_norm)
            for all_row in all_question_to_row.get(key, []):
                all_ws.cell(row=all_row, column=all_answer_col, value=answer)

    wb.save(EXCEL_PATH)
    print(f"Filled {filled} answers from transcript.")
    print(f"Not found in transcript: {not_found}.")


if __name__ == "__main__":
    main()
