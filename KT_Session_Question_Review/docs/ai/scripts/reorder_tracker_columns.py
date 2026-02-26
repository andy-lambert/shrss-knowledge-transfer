#!/usr/bin/env python3
"""Reorder columns in all sheets to: Status, Date Asked, Answered On, Answered By, Asked By, Question/Comment, Answer, Notes."""
from pathlib import Path
import openpyxl

DIR = Path(__file__).resolve().parent
WB_PATH = DIR / "SHRSS_Adobe_KT_Session_Follow_Up_Tracker.xlsx"

# Current order: 1=Question/Comment, 2=Date Asked, 3=Asked By, 4=Answer, 5=Answered On, 6=Answered By, 7=Status, 8=Notes
# New order: 1=Status, 2=Date Asked, 3=Answered On, 4=Answered By, 5=Asked By, 6=Question/Comment, 7=Answer, 8=Notes
OLD_TO_NEW = (7, 2, 5, 6, 3, 1, 4, 8)  # new_col[i] = old_col[OLD_TO_NEW[i]-1], 1-based
NEW_HEADERS = ("Status", "Date Asked", "Answered On", "Answered By", "Asked By", "Question/Comment", "Answer", "Notes")


def reorder_row(old_values):
    """Take 8 values in current order; return tuple in new order."""
    if len(old_values) < 8:
        old_values = list(old_values) + [None] * (8 - len(old_values))
    return tuple(old_values[i - 1] for i in OLD_TO_NEW)


def main():
    wb = openpyxl.load_workbook(WB_PATH)
    for ws in wb.worksheets:
        max_row = ws.max_row
        if max_row < 1:
            continue
        # Read all rows
        rows = []
        for r in range(1, max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            rows.append(row)
        # Reorder and write back
        for r, old_row in enumerate(rows, 1):
            new_row = reorder_row(old_row)
            if r == 1:
                # Header row: use exact new headers
                for c, val in enumerate(NEW_HEADERS, 1):
                    ws.cell(row=r, column=c, value=val)
            else:
                for c, val in enumerate(new_row, 1):
                    ws.cell(row=r, column=c, value=val)
    wb.save(WB_PATH)
    print("Columns reordered in all sheets.")


if __name__ == "__main__":
    main()
