#!/usr/bin/env python3
"""Copy AI-Generated, Confidence, Reasoning from Cursor_AI_Detect into AI_Candidates by matching Question."""
import re
from pathlib import Path

import openpyxl

WB_PATH = Path(__file__).resolve().parent / "SHRSS_KT_Session_Question_Review.xlsx"


def normalize_question(q):
    if q is None:
        return ""
    s = str(q).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def main():
    wb = openpyxl.load_workbook(WB_PATH)
    cd = wb["Cursor_AI_Detect"]
    ac = wb["AI_Candidates"]
    # Build lookup: normalized question -> (AI-Generated, Confidence, Reasoning)
    # Cursor_AI_Detect: col 3 = Question, 4 = AI-Generated, 5 = Confidence, 6 = Reasoning
    lookup = {}
    for r in range(2, cd.max_row + 1):
        q = cd.cell(row=r, column=3).value
        key = normalize_question(q)
        if key not in lookup:  # first occurrence wins
            lookup[key] = (
                cd.cell(row=r, column=4).value,
                cd.cell(row=r, column=5).value,
                cd.cell(row=r, column=6).value,
            )
    # AI_Candidates: col 2 = Question, 4 = AI-Generated, 5 = Confidence, 6 = Reasoning
    matched = 0
    unmatched = []
    for r in range(2, ac.max_row + 1):
        q = ac.cell(row=r, column=2).value
        key = normalize_question(q)
        vals = lookup.get(key)
        if vals is not None:
            ac.cell(row=r, column=4, value=vals[0])
            ac.cell(row=r, column=5, value=vals[1])
            ac.cell(row=r, column=6, value=vals[2])
            matched += 1
        else:
            unmatched.append((r, (str(q)[:60] + "..." if q and len(str(q)) > 60 else q)))
    wb.save(WB_PATH)
    print(f"Matched {matched} of {ac.max_row - 1} rows. Copied AI-Generated, Confidence, Reasoning into AI_Candidates.")
    if unmatched:
        print(f"Unmatched rows: {len(unmatched)}")
        for row, q in unmatched[:10]:
            print(f"  Row {row}: {repr(q)}")
        if len(unmatched) > 10:
            print(f"  ... and {len(unmatched) - 10} more.")


if __name__ == "__main__":
    main()
